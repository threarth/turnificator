"""
tests/test_modello_struttura.py — la struttura turni letta da un foglio Excel.

Chi arriva da un foglio di calcolo ha gia' tutto scritto li'. Il foglio
`Tabelle` del modello e' la struttura in forma leggibile dalla macchina, e le
righe della griglia sono generate proprio da quella: importarla significa
ottenere turni che combaciano riga per riga con il modello, ed e' cio' che
permettera' poi di riesportarci dentro un mese.

I test lavorano su un foglio costruito qui, non sul modello reale: cosi'
dicono cosa il lettore pretende, invece di fotografare un file.
"""

import importlib.util
import io
import os

import openpyxl
import pytest

from tests.conftest import _open_sqlcipher


_PERCORSO = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'app', 'services', 'modello_struttura.py'
)
_spec = importlib.util.spec_from_file_location('modello_struttura', _PERCORSO)
ms = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(ms)


# Le posizioni del modello: nome, tipologia e sigla per ciascuna fascia.
COLONNE = {
    'mattina':    (1, 2, 3),
    'pomeriggio': (5, 6, 7),
    'notte':      (9, 10, 11),
}


def _foglio(turni, persone=(), richieste=(), assenze=()):
    """
    Costruisce un modello minimo in memoria.

    Args:
        turni (dict): fascia → [(nome, tipologia, sigla)].
        persone (iterable): (sigla, cognome, nome).
        richieste (iterable): sigle di tutto cio' che si puo' chiedere.
        assenze (iterable): quali di quelle sono assenze.

    Returns:
        BytesIO: il file .xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tabelle'

    for fascia, righe in turni.items():
        c_nome, c_tipo, c_sigla = COLONNE[fascia]
        for i, (nome, tipologia, sigla) in enumerate(righe):
            ws.cell(2 + i, c_nome, nome)
            ws.cell(2 + i, c_tipo, tipologia)
            ws.cell(2 + i, c_sigla, sigla)

    for i, (sigla, cognome, nome) in enumerate(persone):
        ws.cell(2 + i, 14, sigla)
        ws.cell(2 + i, 15, cognome)
        ws.cell(2 + i, 16, nome)

    for i, sigla in enumerate(richieste):
        ws.cell(2 + i, 18, sigla)
    for i, sigla in enumerate(assenze):
        ws.cell(2 + i, 19, sigla)

    dati = io.BytesIO()
    wb.save(dati)
    dati.seek(0)

    return dati


# ---------------------------------------------------------------------------
# I turni
# ---------------------------------------------------------------------------

def test_i_turni_arrivano_nell_ordine_del_foglio():
    """L'ordine e' quello con cui la griglia dispone le righe: va conservato."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; '), ('S.G. TC', 'TC', 'tcSGm; ')],
        'pomeriggio': [('S.G. DEA 1 P', 'DEA', 'deaP; ')],
    }))

    assert [t['nome'] for t in letto['turni']] == ['S.G. DEA 1', 'S.G. TC', 'S.G. DEA 1 P']
    assert [t['fascia'] for t in letto['turni']] == ['mattina', 'mattina', 'pomeriggio']


def test_le_righe_lasciate_libere_non_diventano_turni():
    """Il foglio riempie con un trattino le righe che non usa."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('-', '-', 'turnovuoto; '), ('S.G. DEA 1', 'DEA', 'deaM; ')],
    }))

    assert [t['nome'] for t in letto['turni']] == ['S.G. DEA 1']


def test_un_turno_senza_sigla_viene_segnalato_non_perso():
    """Senza sigla non so in che struttura metterlo: lo dico."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; '), ('Orfano', 'DEA', '')],
    }))

    assert [t['nome'] for t in letto['turni']] == ['S.G. DEA 1']
    assert any('Orfano' in a for a in letto['avvisi'])


def test_un_foglio_senza_turni_e_un_errore():
    with pytest.raises(ValueError, match='nessun turno'):
        ms.leggi_struttura(_foglio({'mattina': []}))


def test_un_foglio_senza_la_tabella_e_un_errore():
    wb = openpyxl.Workbook()
    wb.active.title = 'Altro'
    dati = io.BytesIO()
    wb.save(dati)
    dati.seek(0)

    with pytest.raises(ValueError, match='Tabelle'):
        ms.leggi_struttura(dati)


# ---------------------------------------------------------------------------
# Le strutture
# ---------------------------------------------------------------------------

def test_la_sede_nel_nome_fa_la_struttura():
    """
    La struttura e' il luogo, e lo dice la prima parola del nome. Due
    metodiche diverse nella stessa sede sono una struttura sola.
    """
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; '), ('S.G. TC', 'TC', 'tcSGm; '),
                    ('ADD. TC', 'TC', 'tcADm; ')],
    }))

    assert [s['nome'] for s in letto['strutture']] == ['S.G.', 'ADD.']


def test_la_punteggiatura_della_sede_non_fa_due_strutture():
    """Nei fogli veri la stessa sede si scrive «ADD.» e «ADD»."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('ADD. TC', 'TC', 'tcADm; '), ('ADD SUPP. TC', 'TC', 'addSupM')],
    }))

    assert len(letto['strutture']) == 1


def test_un_turno_che_non_nomina_il_luogo_segue_il_suo_posto():
    """
    La notte si chiama «NOTTE - 1° med.» e la sede non la dice: la prende
    dagli altri turni dello stesso posto, che e' la sigla senza la fascia.
    """
    letto = ms.leggi_struttura(_foglio({
        'mattina':    [('S.G. DEA 1', 'DEA', 'deaM; ')],
        'pomeriggio': [('S.G. DEA 1 P', 'DEA', 'deaP; ')],
        'notte':      [('NOTTE - 1 med.', 'DEANOTTE', 'deaN')],
    }))

    assert [s['nome'] for s in letto['strutture']] == ['S.G.']
    notte = next(t for t in letto['turni'] if t['fascia'] == 'notte')
    assert notte['struttura_nome'] == 'S.G.'


def test_un_posto_di_cui_nessun_turno_nomina_il_luogo_resta_a_se():
    """Senza nessun luogo scritto, la struttura e' la sigla del posto."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('Ambulatorio', 'TC', 'ambM; ')],
    }))

    assert [s['nome'] for s in letto['strutture']] == ['AMB']


# ---------------------------------------------------------------------------
# Persone, tipologie, tipi richiesta
# ---------------------------------------------------------------------------

def test_le_persone_arrivano_con_sigla_cognome_e_nome():
    letto = ms.leggi_struttura(_foglio(
        {'mattina': [('S.G. DEA 1', 'DEA', 'deaM; ')]},
        persone=[('assae', 'Assael', 'Filippo'), ('CARAV', 'Caravani', 'Francesca')],
    ))

    assert letto['persone'] == [
        {'sigla': 'ASSAE', 'cognome': 'Assael', 'nome': 'Filippo'},
        {'sigla': 'CARAV', 'cognome': 'Caravani', 'nome': 'Francesca'},
    ]


def test_i_segnaposto_non_diventano_persone():
    """Il foglio tiene righe 'Vuoto' e 'chiusa' per le posizioni non coperte."""
    letto = ms.leggi_struttura(_foglio(
        {'mattina': [('S.G. DEA 1', 'DEA', 'deaM; ')]},
        persone=[('ASSAE', 'Assael', 'Filippo'), ('VUOTO', 'Vuoto', 'Vuoto'),
                 ('chiusa', 'chiusa', '')],
    ))

    assert [p['sigla'] for p in letto['persone']] == ['ASSAE']


def test_le_tipologie_sono_quelle_dei_turni_senza_ripetizioni():
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; '), ('S.G. DEA 2', 'DEA', 'deaM; '),
                    ('S.G. TC', 'TC', 'tcSGm; ')],
    }))

    assert letto['tipologie'] == ['DEA', 'TC']


def test_le_assenze_si_distinguono_da_cio_che_si_chiede():
    """
    Il foglio tiene due elenchi: tutto cio' che si puo' chiedere, e quali di
    quelle cose sono un'assenza. Il secondo e' un sottoinsieme del primo.
    """
    letto = ms.leggi_struttura(_foglio(
        {'mattina': [('S.G. DEA 1', 'DEA', 'deaM; ')]},
        richieste=['CO', 'M', 'P', 'N', 'ROMC'],
        assenze=['CO', 'ROMC'],
    ))

    per_sigla = {t['sigla']: t['tipo'] for t in letto['tipi_richiesta']}
    assert per_sigla == {
        'CO': 'assenza', 'ROMC': 'assenza',
        'M': 'lavorativo', 'P': 'lavorativo', 'N': 'lavorativo',
    }


# ---------------------------------------------------------------------------
# Dal foglio al programma
# ---------------------------------------------------------------------------

def _modello_di_prova():
    """Un modello con due strutture, tre turni, due persone."""
    return _foglio(
        {
            'mattina':    [('S.G. DEA 1', 'DEA', 'deaM; '), ('ADD. TC', 'TC', 'tcADm; ')],
            'pomeriggio': [('S.G. DEA 1 P', 'DEA', 'deaP; ')],
        },
        persone=[('ROSSI', 'Rossi', 'Mario'), ('VERDI', 'Verdi', 'Anna')],
        richieste=['CO', 'M', 'P'],
        assenze=['CO'],
    )


def _allega(dati, nome='modello.xlsx'):
    """Il file nella forma che il client di test manda come multipart."""
    dati.seek(0)
    return {'file': (io.BytesIO(dati.read()), nome)}


def test_l_analisi_racconta_senza_scrivere(client, admin_token, auth):
    """Prima si vede cosa verrebbe creato, poi si decide."""
    rv = client.post('/api/admin/modello/analizza',
                     data=_allega(_modello_di_prova()),
                     content_type='multipart/form-data',
                     headers=auth(admin_token))
    assert rv.status_code == 200, rv.get_json()

    corpo = rv.get_json()
    assert [s['nome'] for s in corpo['strutture']] == ['S.G.', 'ADD.']
    assert len(corpo['turni']) == 3
    assert [p['sigla'] for p in corpo['persone']] == ['ROSSI', 'VERDI']

    # Niente e' stato creato.
    presets = client.get('/api/admin/struttura-presets',
                         headers=auth(admin_token)).get_json()
    assert not any(p['nome'] == 'Dal foglio' for p in presets.get('presets', []))


def test_un_file_che_non_e_un_foglio_viene_rifiutato(client, admin_token, auth):
    rv = client.post('/api/admin/modello/analizza',
                     data={'file': (io.BytesIO(b'non sono un xlsx'), 'finto.xlsx')},
                     content_type='multipart/form-data',
                     headers=auth(admin_token))

    assert rv.status_code == 400
    assert 'foglio Excel' in rv.get_json()['errore']


def test_applicare_crea_struttura_tipologie_e_persone(client, admin_token, auth):
    """La struttura del foglio diventa la struttura turni del programma."""
    rv = client.post('/api/admin/modello/applica',
                     data={**_allega(_modello_di_prova()), 'nome_preset': 'Dal foglio'},
                     content_type='multipart/form-data',
                     headers=auth(admin_token))
    assert rv.status_code == 201, rv.get_json()

    corpo = rv.get_json()
    assert corpo['strutture'] == 2
    assert corpo['turni'] == 3
    assert [c['sigla'] for c in corpo['persone_create']] == ['ROSSI', 'VERDI']
    assert all(c['password'] for c in corpo['persone_create'])

    rv = client.get(f"/api/admin/struttura-presets/{corpo['preset_id']}/struttura",
                    headers=auth(admin_token))
    assert rv.status_code == 200, rv.get_json()
    sovragruppi = rv.get_json()['struttura']

    assert [sg['nome'] for sg in sovragruppi] == ['S.G.', 'ADD.']
    # S.G. ha due fasce, ADD. una sola: il gruppo nasce dalla fascia.
    assert [len(sg['gruppi']) for sg in sovragruppi] == [2, 1]
    nomi_turni = [t['nome'] for sg in sovragruppi for g in sg['gruppi'] for t in g['turni']]
    assert sorted(nomi_turni) == ['ADD. TC', 'S.G. DEA 1', 'S.G. DEA 1 P']


def test_chi_c_e_gia_non_viene_ricreato(client, admin_token, auth):
    """Il foglio non sovrascrive le persone del programma."""
    client.post('/api/admin/users',
                json={'username': 'rossi', 'password': 'Password2027',
                      'sigla': 'ROSSI', 'role': 'basic'},
                headers=auth(admin_token))

    rv = client.post('/api/admin/modello/applica',
                     data={**_allega(_modello_di_prova()), 'nome_preset': 'Dal foglio 2'},
                     content_type='multipart/form-data',
                     headers=auth(admin_token))
    assert rv.status_code == 201, rv.get_json()

    assert [c['sigla'] for c in rv.get_json()['persone_create']] == ['VERDI']


def test_il_foglio_resta_nel_tenant(client, admin_token, auth):
    """Diventa il modello in cui riesportare i mesi programmati."""
    assert client.get('/api/admin/modello',
                      headers=auth(admin_token)).get_json()['modello'] is None

    client.post('/api/admin/modello/applica',
                data={**_allega(_modello_di_prova(), 'turni_2026.xlsx'),
                      'nome_preset': 'Dal foglio 3'},
                content_type='multipart/form-data',
                headers=auth(admin_token))

    modello = client.get('/api/admin/modello',
                         headers=auth(admin_token)).get_json()['modello']
    assert modello['nome_file'] == 'turni_2026.xlsx'
    assert modello['byte'] > 0


def test_solo_l_amministratore_importa_una_struttura(client, manager_token, auth):
    rv = client.post('/api/admin/modello/analizza',
                     data=_allega(_modello_di_prova()),
                     content_type='multipart/form-data',
                     headers=auth(manager_token))

    assert rv.status_code == 403


# ---------------------------------------------------------------------------
# Correggere la deduzione
# ---------------------------------------------------------------------------

def test_rinominare_una_struttura_la_ribattezza():
    """La sede dedotta dal nome e' una proposta: si corregge."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; ')],
    }))

    ms.rinomina_strutture(letto, {'SG': 'San Giovanni'})

    assert [s['nome'] for s in letto['strutture']] == ['San Giovanni']
    assert letto['turni'][0]['struttura_nome'] == 'San Giovanni'


def test_due_strutture_chiamate_uguale_si_fondono():
    """E' il modo per dire «questi due sono lo stesso posto»."""
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; '), ('ADD. TC', 'TC', 'tcADm; ')],
    }))
    assert len(letto['strutture']) == 2

    ms.rinomina_strutture(letto, {'SG': 'Unica', 'ADD': 'Unica'})

    assert [s['nome'] for s in letto['strutture']] == ['Unica']
    assert {t['struttura'] for t in letto['turni']} == {'Unica'}


def test_una_struttura_non_nominata_resta_com_e():
    letto = ms.leggi_struttura(_foglio({
        'mattina': [('S.G. DEA 1', 'DEA', 'deaM; '), ('ADD. TC', 'TC', 'tcADm; ')],
    }))

    ms.rinomina_strutture(letto, {'SG': 'San Giovanni'})

    assert [s['nome'] for s in letto['strutture']] == ['San Giovanni', 'ADD.']


def test_le_correzioni_arrivano_fino_alla_struttura_creata(client, admin_token, auth):
    """Quello che l'utente scrive nel riquadro e' quello che viene creato."""
    import json as _json

    rv = client.post('/api/admin/modello/applica',
                     data={**_allega(_modello_di_prova()),
                           'nome_preset': 'Con correzioni',
                           'strutture': _json.dumps({'SG': 'Sede unica',
                                                     'ADD': 'Sede unica'})},
                     content_type='multipart/form-data',
                     headers=auth(admin_token))
    assert rv.status_code == 201, rv.get_json()
    assert rv.get_json()['strutture'] == 1

    rv = client.get(f"/api/admin/struttura-presets/{rv.get_json()['preset_id']}/struttura",
                    headers=auth(admin_token))
    sovragruppi = rv.get_json()['struttura']

    assert [sg['nome'] for sg in sovragruppi] == ['Sede unica']


# ---------------------------------------------------------------------------
# Una struttura sola per organizzazione
# ---------------------------------------------------------------------------

def _strutture_del_tenant(client, admin_token, auth):
    return client.get('/api/admin/struttura-presets',
                      headers=auth(admin_token)).get_json()['presets']


def test_importare_due_volte_non_lascia_due_strutture(client, admin_token, auth):
    """
    L'organizzazione ha una struttura sola: il secondo import sostituisce il
    primo, non gli si affianca.
    """
    prima = len(_strutture_del_tenant(client, admin_token, auth))

    for nome in ('Primo giro', 'Secondo giro'):
        rv = client.post('/api/admin/modello/applica',
                         data={**_allega(_modello_di_prova()), 'nome_preset': nome},
                         content_type='multipart/form-data',
                         headers=auth(admin_token))
        assert rv.status_code == 201, rv.get_json()

    dopo = _strutture_del_tenant(client, admin_token, auth)
    assert len(dopo) == prima
    assert rv.get_json()['sostituita'] is True


def test_la_struttura_importata_diventa_quella_dell_organizzazione(client, admin_token, auth):
    """Senza il segno di predefinita, il calendario non saprebbe quale usare."""
    rv = client.post('/api/admin/modello/applica',
                     data={**_allega(_modello_di_prova()), 'nome_preset': 'Dal foglio'},
                     content_type='multipart/form-data',
                     headers=auth(admin_token))
    preset_id = rv.get_json()['preset_id']

    strutture = _strutture_del_tenant(client, admin_token, auth)
    predefinite = [p for p in strutture if p['is_default']]

    assert [p['id'] for p in predefinite] == [preset_id]


def test_il_calendario_non_chiede_quale_struttura(client, admin_token, auth):
    """Ce n'e' una sola: chiederlo sarebbe una domanda senza alternative."""
    client.post('/api/admin/modello/applica',
                data={**_allega(_modello_di_prova()), 'nome_preset': 'Dal foglio'},
                content_type='multipart/form-data', headers=auth(admin_token))

    rv = client.post('/api/admin/calendari', json={'mese': 3, 'anno': 2028},
                     headers=auth(admin_token))

    assert rv.status_code == 201, rv.get_json()


def test_i_turni_importati_finiscono_nel_calendario(client, admin_token, auth):
    """La struttura del foglio e' quella con cui si costruisce il mese."""
    client.post('/api/admin/modello/applica',
                data={**_allega(_modello_di_prova()), 'nome_preset': 'Dal foglio'},
                content_type='multipart/form-data', headers=auth(admin_token))

    rv = client.post('/api/admin/calendari', json={'mese': 4, 'anno': 2028},
                     headers=auth(admin_token))
    cal_id = rv.get_json()['id']

    rv = client.get(f'/api/manager/calendari/{cal_id}/struttura',
                    headers=auth(admin_token))
    sovragruppi = rv.get_json()['sovragruppi']

    assert [sg['nome'] for sg in sovragruppi] == ['S.G.', 'ADD.']
    nomi = [t['descrizione'] for sg in sovragruppi
            for g in sg['gruppi'] for t in g['turni']]
    assert sorted(nomi) == ['ADD. TC', 'S.G. DEA 1', 'S.G. DEA 1 P']


def test_senza_predefinita_si_usa_l_ultima_creata(client, admin_token, auth, _test_env):
    """
    Un'installazione che ha accumulato strutture prima che la regola
    esistesse non deve restare senza: si prende l'ultima, e il primo import
    la marca come quella dell'organizzazione.
    """
    client.post('/api/admin/struttura-presets', json={'nome': 'Vecchia A'},
                headers=auth(admin_token))
    rv = client.post('/api/admin/struttura-presets', json={'nome': 'Vecchia B'},
                     headers=auth(admin_token))
    ultima = rv.get_json()['id']

    # Nessuna e' predefinita: e' lo stato di un'installazione che ha
    # accumulato strutture prima che la regola esistesse, e dall'API non si
    # puo' produrre — la scelta si sposta, non si toglie.
    db = _open_sqlcipher(_test_env['tenant_path'], _test_env['tenant_key'])
    db.execute("UPDATE struttura_presets SET is_default=0")
    db.commit()

    rv = client.post('/api/admin/modello/applica',
                     data={**_allega(_modello_di_prova()), 'nome_preset': 'Dal foglio'},
                     content_type='multipart/form-data', headers=auth(admin_token))

    assert rv.status_code == 201, rv.get_json()
    assert rv.get_json()['preset_id'] == ultima
    predefinite = [p for p in _strutture_del_tenant(client, admin_token, auth)
                   if p['is_default']]
    assert [p['id'] for p in predefinite] == [ultima]
