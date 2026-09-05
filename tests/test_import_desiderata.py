"""
tests/test_import_desiderata.py — le richieste di un mese, prese da un foglio.

I lavoratori compilano i desiderata in un foglio di calcolo: una riga per
persona, una colonna per giorno. Ridigitarli nella griglia e' un lavoro lungo
su dati che esistono gia'.

L'import **sostituisce il mese intero**: e' la scelta di chi lo usa — il
foglio e' la verita' — e per questo prima si guarda, e con discrepanze fra le
persone del foglio e quelle del programma serve una conferma esplicita.
"""

import datetime
import importlib.util
import io
import os

import openpyxl
import pytest

from tests.conftest import _open_sqlcipher


_PERCORSO = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'app', 'services', 'modello_desiderata.py'
)
_spec = importlib.util.spec_from_file_location('modello_desiderata', _PERCORSO)
md = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(md)


# Le posizioni del modello.
RIGA_DATE = 8
PRIMA_RIGA = 10
PRIMA_COLONNA_GIORNI = 2


def _foglio(mese, anno, righe, giorni=31, dichiara_mese=True, titolo=None):
    """
    Costruisce un foglio di desiderata in memoria.

    Args:
        mese, anno (int): il mese di cui parla il foglio.
        righe (list): [(sigla, {giorno: codice})].
        giorni (int): quante colonne di giorni scrivere.
        dichiara_mese (bool): scrivere le celle MESE e ANNO.
        titolo (str|None): nome del foglio.

    Returns:
        BytesIO: il file .xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titolo or 'MESE'

    per_giorno = {}
    for i in range(giorni):
        colonna = PRIMA_COLONNA_GIORNI + i
        ws.cell(RIGA_DATE, colonna, datetime.datetime(anno, mese, i + 1))
        per_giorno[i + 1] = colonna

    if dichiara_mese:
        colonna = PRIMA_COLONNA_GIORNI + giorni + 1
        ws.cell(RIGA_DATE, colonna, 'MESE')
        ws.cell(RIGA_DATE + 1, colonna, mese)
        ws.cell(RIGA_DATE, colonna + 1, 'ANNO')
        ws.cell(RIGA_DATE + 1, colonna + 1, anno)

    for i, (sigla, richieste) in enumerate(righe):
        ws.cell(PRIMA_RIGA + i, 1, sigla)
        for giorno, codice in richieste.items():
            ws.cell(PRIMA_RIGA + i, per_giorno[giorno], codice)

    dati = io.BytesIO()
    wb.save(dati)
    dati.seek(0)

    return dati


# ---------------------------------------------------------------------------
# Leggere il foglio
# ---------------------------------------------------------------------------

def test_il_mese_lo_dice_il_foglio_non_il_suo_nome():
    """Il nome di un foglio si rinomina; le celle che dichiarano il mese no."""
    letto = md.leggi_desiderata(_foglio(10, 2026, [('ROSSI', {1: 'M'})],
                                        titolo='UN NOME QUALSIASI'))

    assert (letto['mese'], letto['anno']) == (10, 2026)


def test_senza_le_celle_del_mese_lo_dicono_le_date():
    letto = md.leggi_desiderata(_foglio(3, 2027, [('ROSSI', {1: 'M'})],
                                        dichiara_mese=False))

    assert (letto['mese'], letto['anno']) == (3, 2027)


def test_le_richieste_arrivano_con_persona_giorno_e_codice():
    letto = md.leggi_desiderata(_foglio(10, 2026, [
        ('ROSSI', {1: 'M', 5: 'CO'}),
        ('VERDI', {2: 'N'}),
    ]))

    assert sorted((r['sigla'], r['giorno'], r['codice']) for r in letto['richieste']) == [
        ('ROSSI', 1, 'M'), ('ROSSI', 5, 'CO'), ('VERDI', 2, 'N'),
    ]


def test_le_celle_vuote_non_sono_richieste():
    letto = md.leggi_desiderata(_foglio(10, 2026, [('ROSSI', {1: 'M', 2: ''})]))

    assert len(letto['richieste']) == 1


def test_i_segnaposto_non_sono_persone():
    """Il foglio tiene righe 'Vuoto', 'chiusa', 'VUOTO_C' per le posizioni libere."""
    letto = md.leggi_desiderata(_foglio(10, 2026, [
        ('ROSSI', {1: 'M'}), ('Vuoto', {1: 'M'}),
        ('chiusa', {}), ('VUOTO_C', {1: 'P'}),
    ]))

    assert letto['persone'] == ['ROSSI']


def test_sotto_il_primo_buco_non_si_leggono_persone():
    """Piu' in basso il foglio tiene piedi di pagina, non lavoratori."""
    dati = _foglio(10, 2026, [('ROSSI', {1: 'M'})])
    wb = openpyxl.load_workbook(dati)
    wb.active.cell(40, 1, 'NOME FILE')
    rifatto = io.BytesIO()
    wb.save(rifatto)
    rifatto.seek(0)

    assert md.leggi_desiderata(rifatto)['persone'] == ['ROSSI']


def test_un_foglio_senza_date_e_un_errore():
    wb = openpyxl.Workbook()
    wb.active.cell(10, 1, 'ROSSI')
    dati = io.BytesIO()
    wb.save(dati)
    dati.seek(0)

    with pytest.raises(ValueError, match='non contiene date'):
        md.leggi_desiderata(dati)


def test_un_foglio_senza_lavoratori_e_un_errore():
    with pytest.raises(ValueError, match='nessun lavoratore'):
        md.leggi_desiderata(_foglio(10, 2026, []))


# ---------------------------------------------------------------------------
# Guardare prima di scrivere
# ---------------------------------------------------------------------------

def _allega(dati, nome='desiderata.xlsx'):
    dati.seek(0)
    return {'file': (io.BytesIO(dati.read()), nome)}


def _crea_calendario(client, token, auth, mese, anno):
    rv = client.post('/api/admin/calendari', json={'mese': mese, 'anno': anno},
                     headers=auth(token))
    assert rv.status_code == 201, rv.get_json()
    return rv.get_json()['id']


def _foglio_del_tenant(mese=5, anno=2029):
    """Un foglio con le sigle e i codici che il tenant di prova conosce."""
    return _foglio(mese, anno, [
        ('BSC', {1: 'M', 2: 'CO', 3: 'M'}),
        ('MGR', {1: 'CO'}),
    ])


def test_l_analisi_dice_di_che_mese_parla_senza_scrivere(client, admin_token, auth):
    rv = client.post('/api/admin/desiderata/analizza',
                     data=_allega(_foglio_del_tenant()),
                     content_type='multipart/form-data', headers=auth(admin_token))
    assert rv.status_code == 200, rv.get_json()

    corpo = rv.get_json()
    assert (corpo['mese'], corpo['anno']) == (5, 2029)
    assert corpo['richieste_nel_foglio'] == 4
    assert corpo['calendario'] is None, 'quel mese non esiste ancora'


def test_l_analisi_segnala_chi_non_si_riconosce(client, admin_token, auth):
    """
    Prima di sostituire un mese bisogna essere certi di parlare delle stesse
    persone: una sigla ignota e' una colonna di richieste che andrebbe persa.
    """
    rv = client.post('/api/admin/desiderata/analizza',
                     data=_allega(_foglio(5, 2029, [('BSC', {1: 'M'}),
                                                    ('IGNOTO', {1: 'M'})])),
                     content_type='multipart/form-data', headers=auth(admin_token))

    corpo = rv.get_json()
    assert corpo['persone_sconosciute'] == ['IGNOTO']
    assert 'BSC' not in corpo['persone_senza_riga']
    assert corpo['discrepanze'] is True


def test_l_analisi_segnala_chi_nel_foglio_non_c_e(client, admin_token, auth):
    rv = client.post('/api/admin/desiderata/analizza',
                     data=_allega(_foglio(5, 2029, [('BSC', {1: 'M'})])),
                     content_type='multipart/form-data', headers=auth(admin_token))

    assert 'MGR' in rv.get_json()['persone_senza_riga']


def test_l_analisi_segnala_le_sigle_di_richiesta_ignote(client, admin_token, auth):
    rv = client.post('/api/admin/desiderata/analizza',
                     data=_allega(_foglio(5, 2029, [('BSC', {1: 'ZZZ'})])),
                     content_type='multipart/form-data', headers=auth(admin_token))

    corpo = rv.get_json()
    assert corpo['codici_sconosciuti'] == ['ZZZ']
    assert corpo['richieste_importabili'] == 0


# ---------------------------------------------------------------------------
# Importare
# ---------------------------------------------------------------------------

def test_senza_calendario_l_import_si_ferma(client, admin_token, auth):
    """Le richieste appartengono a un mese: senza, non c'e' dove metterle."""
    rv = client.post('/api/admin/desiderata/importa',
                     data={**_allega(_foglio_del_tenant(mese=7, anno=2029)),
                           'conferma': 'true'},
                     content_type='multipart/form-data', headers=auth(admin_token))

    assert rv.status_code == 409
    assert 'calendario' in rv.get_json()['errore']


def test_con_discrepanze_serve_una_conferma(client, admin_token, auth):
    _crea_calendario(client, admin_token, auth, 5, 2029)

    rv = client.post('/api/admin/desiderata/importa',
                     data=_allega(_foglio(5, 2029, [('BSC', {1: 'M'})])),
                     content_type='multipart/form-data', headers=auth(admin_token))

    assert rv.status_code == 409
    assert rv.get_json()['codice'] == 'discrepanze'


def test_confermando_le_richieste_entrano(client, admin_token, auth, basic_token):
    cal_id = _crea_calendario(client, admin_token, auth, 5, 2029)

    rv = client.post('/api/admin/desiderata/importa',
                     data={**_allega(_foglio_del_tenant()), 'conferma': 'true'},
                     content_type='multipart/form-data', headers=auth(admin_token))
    assert rv.status_code == 200, rv.get_json()
    assert rv.get_json()['importate'] == 4

    rv = client.get(f'/api/basic/calendari/{cal_id}/desiderata',
                    headers=auth(basic_token))
    assert rv.status_code == 200, rv.get_json()
    miei = {d['giorno']: d['req_sigla'] for d in rv.get_json()['desiderata']}
    assert miei == {1: 'M', 2: 'CO', 3: 'M'}


def test_l_import_sostituisce_il_mese(client, admin_token, auth, _test_env):
    """Il foglio e' la verita': quello che c'era prima non resta accanto."""
    cal_id = _crea_calendario(client, admin_token, auth, 5, 2029)

    client.post('/api/admin/desiderata/importa',
                data={**_allega(_foglio(5, 2029, [('BSC', {10: 'M'}), ('MGR', {11: 'M'})])),
                      'conferma': 'true'},
                content_type='multipart/form-data', headers=auth(admin_token))

    rv = client.post('/api/admin/desiderata/importa',
                     data={**_allega(_foglio_del_tenant()), 'conferma': 'true'},
                     content_type='multipart/form-data', headers=auth(admin_token))
    assert rv.status_code == 200, rv.get_json()
    assert rv.get_json()['sostituite'] == 2

    db = _open_sqlcipher(_test_env['tenant_path'], _test_env['tenant_key'])
    giorni = [r['giorno'] for r in db.execute(
        "SELECT giorno FROM desiderata WHERE calendario_id=? ORDER BY giorno", (cal_id,)
    )]

    assert giorni == [1, 1, 2, 3], 'i giorni 10 e 11 non devono sopravvivere'


def test_quello_che_non_si_riconosce_resta_fuori_e_si_conta(client, admin_token, auth):
    _crea_calendario(client, admin_token, auth, 5, 2029)

    rv = client.post('/api/admin/desiderata/importa',
                     data={**_allega(_foglio(5, 2029, [('BSC', {1: 'M', 2: 'ZZZ'}),
                                                       ('IGNOTO', {1: 'M'})])),
                           'conferma': 'true'},
                     content_type='multipart/form-data', headers=auth(admin_token))

    corpo = rv.get_json()
    assert corpo['importate'] == 1
    assert corpo['saltate'] == 2


def test_se_i_desiderata_erano_congelati_la_copia_si_rifa(client, admin_token, auth,
                                                          manager_token, _test_env):
    """
    Pianificare su una copia di lavoro che nessuno ha piu' chiesto sarebbe
    peggio che non importare.
    """
    cal_id = _crea_calendario(client, admin_token, auth, 5, 2029)
    rv = client.post(f'/api/admin/calendari/{cal_id}/congela',
                     headers=auth(manager_token))
    assert rv.status_code in (200, 201), rv.get_json()

    rv = client.post('/api/admin/desiderata/importa',
                     data={**_allega(_foglio_del_tenant()), 'conferma': 'true'},
                     content_type='multipart/form-data', headers=auth(admin_token))
    assert rv.status_code == 200, rv.get_json()
    assert rv.get_json()['copia_di_lavoro_rifatta'] is True

    db = _open_sqlcipher(_test_env['tenant_path'], _test_env['tenant_key'])
    n = db.execute("SELECT COUNT(*) AS n FROM working_desiderata WHERE calendario_id=?",
                   (cal_id,)).fetchone()['n']
    assert n == 4


def test_solo_l_amministratore_importa_i_desiderata(client, manager_token, auth):
    rv = client.post('/api/admin/desiderata/analizza',
                     data=_allega(_foglio_del_tenant()),
                     content_type='multipart/form-data', headers=auth(manager_token))

    assert rv.status_code == 403
