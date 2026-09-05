"""
genera_config_turni.py — la semantica del modello turni in forma tabellare.

Nuova feature. Il modello Excel (`modello_turni_set26.xlsx`) tiene la propria
semantica in tre posti che non si possono interrogare: dentro le sigle
(`deaM;` = DEA + mattina), dentro la geometria delle righe (una riga e'
"mattina" perche' sta fra la 22 e la 39) e dentro le formule del Riepilogo
(`COUNTIF(...,"*tcSG*")`). Questo script la estrae e la riscrive come tabelle
normalizzate: una riga per entita', una colonna per attributo.

Il risultato e' una cartella di lavoro nuova, non una modifica dell'originale.
La ragione e' tecnica: openpyxl non conosce le estensioni x14 e, riscrivendo il
file di partenza, ne cancellerebbe le 38 formattazioni condizionali e le
validazioni dati. L'originale viene aperto in sola lettura.

Le tabelle sono Tabelle Excel vere (ListObject): si allungano da sole, e le
tendine che le citano restano agganciate quando si aggiunge una riga.
"""

import calendar
import datetime
import re
import unicodedata

import openpyxl
import openpyxl.formatting.rule
import openpyxl.styles.differential
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


MODELLO_ORIGINE = 'modello_turni_set26.xlsx'
FILE_USCITA = 'modello_config.xlsx'
FOGLIO_TABELLE = 'Tabelle'

# Le sedi: il presidio fisico incrociato con la tipologia di turno.
# Un turno aggiuntivo dell'Addolorata sta all'Addolorata come luogo, ma il
# solver lo valuta e lo conta a parte — percio' e' una sede sua.
# I turni di una sede aggiuntiva stanno fuori dalla contabilita' ordinaria:
# non concorrono al monte turni dovuti e si retribuiscono a parte.
# (codice, nome, presidio, tipologia, conta_nei_dovuti,
#  pagamento_separato, descrizione)
SEDI = (
    ('S.G.',     'San Giovanni',            'S.G.', 'ordinario',
     'SI', 'NO', 'Ospedale principale, sede del DEA'),
    ('ADDO',     'Addolorata',              'ADDO', 'ordinario',
     'SI', 'NO', 'Ambulatori e oncologia'),
    ('S.M.',     'Santa Maria',             'S.M.', 'ordinario',
     'SI', 'NO', 'Mammografia, ecografia e biopsia mammaria'),
    ('aggiuSG',  'Aggiuntiva San Giovanni', 'S.G.', 'aggiuntiva',
     'NO', 'SI', 'Turni aggiuntivi al San Giovanni'),
    ('aggiuAdd', 'Aggiuntiva Addolorata',   'ADDO', 'aggiuntiva',
     'NO', 'SI', 'Turni aggiuntivi all Addolorata'),
    ('aggiuSM',  'Aggiuntiva Santa Maria',  'S.M.', 'aggiuntiva',
     'NO', 'SI', 'Turni aggiuntivi a Santa Maria'),
)

# I presidi si ricavano dalle sedi: sono il luogo fisico, senza la
# distinzione fra ordinario e aggiuntivo.
PRESIDI = (
    ('S.G.', 'San Giovanni', 'Ospedale principale, sede del DEA'),
    ('ADDO', 'Addolorata',   'Ambulatori e oncologia'),
    ('S.M.', 'Santa Maria',  'Polo senologico'),
)
TIPOLOGIE_SEDE = ('ordinario', 'aggiuntiva')

# Le fasce orarie, dai default di migrations/init_db.sql (flag_turno).
# Il peso e' la durata netta rapportata al turno tipo di 380 minuti.
FASCE = (
    ('mattina',    'diurno',   '08:00', '14:20', 10, 1.0, 'NO'),
    ('pomeriggio', 'diurno',   '14:00', '20:20', 10, 1.0, 'NO'),
    ('lunga',      'diurno',   '08:00', '20:40', 10, 2.0, 'SI'),
    ('notte',      'notturno', '20:00', '08:40', 10, 2.0, 'NO'),
)

# I blocchi di turni dentro Tabelle: fascia, colonna del nome, prima e
# ultima riga, colonna della sigla, se sono turni aggiuntivi.
BLOCCHI = (
    ('mattina',    'A',  2, 19, 'C', False),
    ('pomeriggio', 'E',  2, 19, 'G', False),
    ('notte',      'I',  2,  3, 'K', False),
    ('mattina',    'A', 27, 34, 'C', True),
    ('pomeriggio', 'E', 27, 34, 'G', True),
)

COLONNE_PERSONE = ('N', 'O', 'P')
RIGHE_PERSONE = (2, 42)
BLOCCO_RICHIESTE = ('R', 2, 13)
BLOCCO_ASSENZE = ('S', 2, 9)
BLOCCO_ASSENZE_FESTIVI = ('T', 2, 2)

# Le righe che il foglio lascia libere per tenere fissa la geometria.
SEGNAPOSTO = ('-', '', 'vuoto', 'vuoto_c', 'vuoto_d', 'vuoto_e',
              'chiusa', 'chiuso')

# La sede si legge dal prefisso del nome. Dove il nome non la dice, la
# dichiara il reparto: le notti sono DEA al San Giovanni, la mammografia
# aggiuntiva sta al polo senologico, l'aggiuntiva generica all'Addolorata.
PREFISSI_SEDE = (
    ('AGGIUNT ADDO',  'aggiuAdd'),
    ('AGGIUNT SG',    'aggiuSG'),
    ('AGGIUNT MAMMO', 'aggiuSM'),
    ('AGGIUNT ALTRO', 'aggiuAdd'),
    ('S.G.',          'S.G.'),
    ('ADD',           'ADDO'),
    ('S.M.',          'S.M.'),
    ('NOTTE',         'S.G.'),
)

# `DEANOTTE` impasta metodica e fascia: separati gli assi, resta DEA.
METODICHE_DA_NORMALIZZARE = {'DEANOTTE': 'DEA'}

# Le metodiche, con la lettura per esteso dove e' ricavabile dall'acronimo.
DESCRIZIONI_METODICHE = {
    'DEA':     'Dipartimento Emergenza e Accettazione',
    'TC':      'Tomografia computerizzata',
    'RM':      'Risonanza magnetica',
    'ECO':     'Ecografia',
    'ECORX':   'Ecografia e radiologia tradizionale',
    'ECODOP':  'Ecocolordoppler',
    'ECOMX':   'Ecografia e mammografia',
    'MX':      'Mammografia',
    'MXBIO':   'Mammografia e biopsia',
    'VABBCEM': 'Biopsia VABB e mammografia con contrasto',
    'SENO':    'Senologia',
    'ALTRO':   'Altro',
}

# I turni che coprono l'emergenza girano sette giorni su sette; gli altri
# seguono il calendario del foglio, che chiude la sola domenica
# (WEEKDAY()=1 alla riga 6, NETWORKDAYS.INTL(...,11,...) in AE14).
METODICHE_SETTE_SU_SETTE = ('DEA',)
GIORNI_CONTINUI = 'LMMGVSD'
GIORNI_FERIALI = 'LMMGVS-'
NECESSITA_DEFAULT = 50
NECESSITA_AGGIUNTIVA = 25
NECESSITA_MASSIMA = 100

# Quanto valgono le preferenze morbide, in centesimi di turno.
PESO_EVITA = -30
PESO_PREFERISCI = 30

# Quante regole per persona prevede la tabella delle preferenze.
REGOLE_PER_PERSONA = 10

# Righe libere lasciate pronte nella tabella dei tetti mensili.
RIGHE_TETTI_LIBERE = 20

# Fin dove arriva un elenco che alimenta una tendina. Serve un limite:
# un riferimento a colonna intera rallenta il foglio in modo grave.
RIGHE_MASSIME_ELENCO = 500
MODI_REGOLA = 'mai,solo,evita,preferisci'
SI_NO = 'SI,NO'

# Il separatore fra nome del turno e fascia nell'etichetta univoca.
SEPARATORE_ETICHETTA = ' · '

# Come si scrive il livello davanti al bersaglio: `sede: S.G.`.
SEPARATORE_LIVELLO = ': '

PASSO_ORDINE = 10
STILE_TABELLA = 'TableStyleMedium2'
LARGHEZZA_MASSIMA = 42


def _testo(cella):
    """Il contenuto di una cella come testo pulito, o stringa vuota."""
    if cella is None or cella.value is None:
        return ''

    return str(cella.value).strip()


def _e_segnaposto(testo):
    """La riga e' una di quelle lasciate libere per riempire la griglia."""
    return testo.lower().strip(' .;') in SEGNAPOSTO


def _sede_di(nome):
    """La sede di un turno, dal prefisso del nome; vuota se non lo dice."""
    maiuscolo = nome.upper()
    for prefisso, sede in PREFISSI_SEDE:
        if maiuscolo.startswith(prefisso):
            return sede

    return ''


def _postazione_di(nome):
    """
    Il posto di lavoro dietro un turno, ripulito del suffisso di fascia.

    Lo stesso posto compare due volte nel foglio, una per la mattina e una
    per il pomeriggio, e non sempre con lo stesso nome: `S.G. TC MOB/128 M`
    e `S.G. TC MOB/128 P` sono la stessa TC. Qui si riconducono a un nome
    solo, correggendo anche le sviste dell'originale (`ADD SUPP. TC P` senza
    punto, `ADD. RM SieM.` con la M maiuscola).

    Args:
        nome (str): nome del turno come sta nel foglio.

    Returns:
        str: nome della postazione.
    """
    pulito = re.sub(r'\s+', ' ', nome).strip()
    pulito = re.sub(r'\s+(M|P)$', '', pulito)
    pulito = re.sub(r'^ADD\.?\s+', 'ADD. ', pulito)

    return pulito


def leggi_turni(ws):
    """
    I turni del foglio Tabelle, con gli assi resi espliciti.

    Ogni turno diventa una riga con sede, metodica, fascia e postazione in
    colonne separate, al posto della sigla che oggi le impasta tutte.

    Args:
        ws: foglio `Tabelle`, a valori.

    Returns:
        list: dizionari, uno per turno, nell'ordine del foglio.
    """
    turni = []
    for fascia, col_nome, prima, ultima, col_sigla, aggiuntiva in BLOCCHI:
        col_tipo = get_column_letter(column_index_from_string(col_nome) + 1)

        for riga in range(prima, ultima + 1):
            nome = _testo(ws[f'{col_nome}{riga}'])
            if _e_segnaposto(nome):
                continue

            metodica = _testo(ws[f'{col_tipo}{riga}'])
            metodica = METODICHE_DA_NORMALIZZARE.get(metodica, metodica)
            sigla = _testo(ws[f'{col_sigla}{riga}']).strip('; ')
            spento = _e_segnaposto(metodica) or _e_segnaposto(sigla)

            turni.append({
                'nome_visualizzato': nome,
                'postazione': _postazione_di(nome),
                'sede': _sede_di(nome),
                'metodica': '' if _e_segnaposto(metodica) else metodica,
                'fascia': fascia,
                # Non diventa una colonna: che il turno sia aggiuntivo lo
                # dice gia' la sua sede, tramite T_Sedi.tipologia.
                '_aggiuntiva': aggiuntiva,
                'sigla': '' if _e_segnaposto(sigla) else sigla,
                'attivo': 'NO' if spento else 'SI',
            })

    return turni


def unifica_postazioni(turni):
    """
    Riconduce a una scrittura sola le postazioni che differiscono per
    maiuscole, e assegna a ogni turno l'etichetta postazione + fascia.

    Nell'originale la stessa RM dell'Addolorata compare come `ADD. RM Siem.`
    la mattina e `ADD. RM SieM.` il pomeriggio. Senza questo passaggio
    resterebbero due postazioni distinte, e una regola scritta sull'una non
    varrebbe per l'altra.

    Args:
        turni (list): i turni come li restituisce `leggi_turni`.

    Returns:
        list: gli stessi turni, con la postazione uniformata.
    """
    canoniche = {}
    for turno in turni:
        chiave = turno['postazione'].upper()
        canoniche.setdefault(chiave, turno['postazione'])
        turno['postazione'] = canoniche[chiave]
        turno['etichetta'] = (f"{turno['postazione']}"
                              f"{SEPARATORE_ETICHETTA}{turno['fascia']}")

    return turni


def completa_turni(turni):
    """
    Aggiunge ai turni i parametri di riempimento, con valori di partenza.

    Questi valori il foglio originale non li contiene: nessuno vi ha mai
    scritto quali turni siano obbligatori o in che giorni aprano. Sono
    quindi proposte da rivedere, non dati letti.

    Args:
        turni (list): i turni come li restituisce `leggi_turni`.

    Returns:
        list: gli stessi turni, con le colonne di riempimento aggiunte.
    """
    for indice, turno in enumerate(turni, start=1):
        emergenza = turno['metodica'] in METODICHE_SETTE_SU_SETTE
        spento = turno['attivo'] == 'NO'

        turno['id'] = f'T{indice:02d}'
        turno['riempimento'] = ('chiuso' if spento
                                else 'obbligatorio' if emergenza
                                else 'opzionale')
        turno['necessita'] = (0 if spento
                              else NECESSITA_MASSIMA if emergenza
                              else NECESSITA_AGGIUNTIVA
                              if turno['_aggiuntiva']
                              else NECESSITA_DEFAULT)
        turno['giorni'] = GIORNI_CONTINUI if emergenza else GIORNI_FERIALI
        turno['festivi'] = 'SI' if emergenza else 'NO'
        turno['superfestivi'] = 'SI' if emergenza else 'NO'
        turno['ordine'] = indice * PASSO_ORDINE

    return turni


def leggi_persone(ws):
    """
    Le persone del foglio Tabelle, senza le righe segnaposto.

    Args:
        ws: foglio `Tabelle`, a valori.

    Returns:
        list: dizionari con acronimo, cognome e nome.
    """
    col_acronimo, col_cognome, col_nome = COLONNE_PERSONE
    prima, ultima = RIGHE_PERSONE

    persone = []
    for riga in range(prima, ultima + 1):
        acronimo = _testo(ws[f'{col_acronimo}{riga}'])
        if _e_segnaposto(acronimo):
            continue

        persone.append({
            'acronimo': acronimo,
            'cognome': _testo(ws[f'{col_cognome}{riga}']),
            'nome': _testo(ws[f'{col_nome}{riga}']),
        })

    return persone


def leggi_colonna(ws, blocco):
    """I valori non vuoti di una colonna, nell'ordine del foglio."""
    colonna, prima, ultima = blocco

    valori = []
    for riga in range(prima, ultima + 1):
        valore = _testo(ws[f'{colonna}{riga}'])
        if valore and not _e_segnaposto(valore) and valore not in valori:
            valori.append(valore)

    return valori


def costruisci_bersagli(turni):
    """
    L'elenco unico dei bersagli su cui si puo' scrivere una regola.

    Ogni voce porta scritto il proprio livello davanti ai due punti —
    `sede: S.G.`, `fascia: notte`, `turno: ADD. TC · mattina`. Serve a tre
    cose: la tendina si autodescrive, chi legge il solver ricava il livello
    dall'etichetta senza cercarlo in tabella, e due livelli possono usare lo
    stesso nome senza ambiguita' (`sede: DEA` e `metodica: DEA` convivono).

    L'ordine e' dal generale al particolare: prima le fasce, poi le sedi, poi
    le metodiche, infine postazioni e singoli turni. La tendina lo rispetta,
    perche' segue l'ordine della tabella.

    Args:
        turni (list): i turni gia' completati.

    Returns:
        tuple: (righe della tabella, elenco dei doppioni trovati).
    """
    voci = [(f[0], 'fascia', f'Tutti i turni di fascia {f[0]}')
            for f in FASCE]

    voci += [(codice, 'tipologia', f'Tutti i turni di tipo {codice}')
             for codice in TIPOLOGIE_SEDE]

    voci += [(sede[0], 'sede', f'{sede[1]} — {sede[-1]}') for sede in SEDI]

    voci += [(metodica, 'metodica', DESCRIZIONI_METODICHE.get(metodica, ''))
             for metodica in sorted({t['metodica'] for t in turni
                                     if t['metodica']})]

    # Il bersaglio di livello turno e' la singola fascia: servono regole
    # che distinguano il DEA 1 della mattina da quello del pomeriggio.
    voci += [(turno['etichetta'], 'turno',
              f"{turno['sede']} · {turno['metodica']} · {turno['fascia']}")
             for turno in turni]

    righe = [[f'{livello}{SEPARATORE_LIVELLO}{nome}', livello, nome, note]
             for nome, livello, note in voci]

    return righe, _doppioni(righe)


def _doppioni(righe):
    """I bersagli che comparirebbero due volte nella tendina."""
    visti = set()
    doppioni = []

    for bersaglio, _, _, _ in righe:
        chiave = unicodedata.normalize('NFKD', bersaglio.upper())
        if chiave in visti:
            doppioni.append(bersaglio)
        visti.add(chiave)

    return doppioni


def costruisci_richieste(ws):
    """
    I tipi di richiesta, distinguendo le assenze dalle preferenze di fascia.

    Args:
        ws: foglio `Tabelle`, a valori.

    Returns:
        list: righe con sigla, tipo, se conta le ore, se vale nei festivi.
    """
    tutte = leggi_colonna(ws, BLOCCO_RICHIESTE)
    assenze = leggi_colonna(ws, BLOCCO_ASSENZE)
    anche_festivi = leggi_colonna(ws, BLOCCO_ASSENZE_FESTIVI)

    righe = []
    for sigla in tutte:
        e_assenza = sigla in assenze
        righe.append([
            sigla,
            'assenza' if e_assenza else 'lavorativo',
            FASCIA_RICHIESTA.get(sigla.upper(), ''),
            # ROMC e' l'unico riposo che non concorre al monte ore.
            'NO' if sigla.upper() == 'ROMC' else 'SI',
            'SI' if sigla in anche_festivi else 'NO',
            '',
        ])

    return righe


def scrivi_tabella(wb, titolo, nome_tabella, intestazioni, righe):
    """
    Crea un foglio con una Tabella Excel vera (ListObject).

    Una Tabella, a differenza di un intervallo, si allunga da sola quando si
    aggiunge una riga in fondo: le tendine e le formule che la citano non
    vanno riagganciate a mano.

    Args:
        wb: cartella di lavoro di destinazione.
        titolo (str): nome del foglio.
        nome_tabella (str): nome della Tabella, es. `T_Turni`.
        intestazioni (list): nomi delle colonne.
        righe (list): liste di valori, una per riga.

    Returns:
        Il foglio creato.
    """
    ws = wb.create_sheet(titolo)
    ws.append(intestazioni)
    for riga in righe:
        ws.append(riga)

    # Una Tabella Excel non puo' essere vuota: senza dati si tiene comunque
    # una riga libera, altrimenti Excel rifiuta di aprire il file.
    ultima = max(len(righe) + 1, 2)
    tabella = Table(displayName=nome_tabella,
                    ref=f'A1:{get_column_letter(len(intestazioni))}{ultima}')
    tabella.tableStyleInfo = TableStyleInfo(
        name=STILE_TABELLA, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tabella)

    _adatta_larghezze(ws, intestazioni, righe)
    ws.freeze_panes = 'A2'

    return ws


def _adatta_larghezze(ws, intestazioni, righe):
    """Allarga le colonne quanto basta al contenuto piu' lungo."""
    for indice, intestazione in enumerate(intestazioni):
        lunghezze = [len(str(intestazione)) + 4]
        lunghezze += [len(str(riga[indice])) + 2 for riga in righe
                      if indice < len(riga) and riga[indice] is not None]

        larghezza = min(max(lunghezze), LARGHEZZA_MASSIMA)
        ws.column_dimensions[get_column_letter(indice + 1)].width = larghezza


def aggiungi_tendina(ws, colonne, sorgente, prima_riga, ultima_riga):
    """
    Aggancia un menu' a tendina a una o piu' colonne di un foglio.

    Args:
        ws: il foglio.
        colonne (list): lettere delle colonne da vincolare.
        sorgente (str): formula dell'elenco, es. `=elenco_bersagli`.
        prima_riga (int): prima riga di dati.
        ultima_riga (int): ultima riga da vincolare.
    """
    # Excel scrive la sorgente senza l'uguale davanti — si vede nelle
    # validazioni del modello originale — e con l'uguale la tendina non
    # si apre. Lo si toglie qui invece che a ogni chiamata.
    validazione = DataValidation(
        type='list', formula1=sorgente.lstrip('='), allowBlank=True,
        showErrorMessage=True, errorTitle='Valore non ammesso',
        error='Scegli una voce dal menu a tendina.')
    ws.add_data_validation(validazione)

    for colonna in colonne:
        validazione.add(f'{colonna}{prima_riga}:{colonna}{ultima_riga}')


def _intervallo_dinamico(foglio):
    """
    Un riferimento che copre la prima colonna di un foglio, intestazione
    esclusa, e si allunga da solo quando si aggiungono righe.

    Si potrebbe scrivere `T_Tabella[colonna]`, che e' piu' leggibile, ma
    non tutte le versioni di Excel accettano un riferimento strutturato
    come sorgente di una validazione.

    La forma con INDEX non e' volatile, a differenza di OFFSET: con le
    tendine appese a millecinquecento celle, un nome volatile fa
    ricalcolare il foglio a ogni battuta e lo rende inusabile.

    Args:
        foglio (str): nome del foglio.

    Returns:
        str: la formula del nome definito.
    """
    return (f"{foglio}!$A$2:INDEX({foglio}!$A$2:$A${RIGHE_MASSIME_ELENCO},"
            f"MAX(1,COUNTA({foglio}!$A$2:$A${RIGHE_MASSIME_ELENCO})))")


def registra_nome(wb, nome, riferimento):
    """Dichiara un nome definito utilizzabile come sorgente di tendina."""
    wb.defined_names[nome] = DefinedName(nome, attr_text=riferimento)


def scrivi_preferenze(wb, persone):
    """
    La tabella delle regole per lavoratore: una riga a testa.

    Ogni regola occupa due colonne, bersaglio e modo, con la tendina su
    entrambe. Le colonne appaiate al posto di una stringa da spacchettare
    (`notte,mai; S.M.,mai`) servono a due cose: la tendina impedisce il
    refuso, e rinominando un turno non resta un riferimento morto dentro
    una cella di testo.

    Args:
        wb: cartella di lavoro di destinazione.
        persone (list): le persone, per precompilare la prima colonna.

    Returns:
        Il foglio creato.
    """
    intestazioni = ['persona']
    for numero in range(1, REGOLE_PER_PERSONA + 1):
        intestazioni += [f'r{numero}_bersaglio', f'r{numero}_modo']
    intestazioni.append('note')

    righe = [[p['acronimo']] + [''] * (REGOLE_PER_PERSONA * 2 + 1)
             for p in persone]
    ws = scrivi_tabella(wb, 'Preferenze', 'T_Preferenze', intestazioni, righe)

    ultima = len(righe) + 1
    colonne_bersaglio = [get_column_letter(2 + indice * 2)
                         for indice in range(REGOLE_PER_PERSONA)]
    colonne_modo = [get_column_letter(3 + indice * 2)
                    for indice in range(REGOLE_PER_PERSONA)]

    aggiungi_tendina(ws, colonne_bersaglio, '=elenco_bersagli', 2, ultima)
    aggiungi_tendina(ws, colonne_modo, f'"{MODI_REGOLA}"', 2, ultima)

    return ws


def scrivi_leggimi(wb, note):
    """Il foglio che dice cosa e' stato letto e cosa e' stato proposto."""
    ws = wb.create_sheet('Leggimi', 0)
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 96

    for riga, (voce, testo) in enumerate(note, start=1):
        ws.cell(riga, 1, voce).font = openpyxl.styles.Font(bold=True)
        ws.cell(riga, 2, testo).alignment = \
            openpyxl.styles.Alignment(wrap_text=True, vertical='top')

    return ws


# Le regole assolute, oggi sepolte nelle formattazioni condizionali di
# `Inserimento`. `*` vale per qualsiasi fascia.
REGOLE = (
    ('notte piu altro turno', 'incompatibilita', 'notte', '*', 0, 'SI',
     'Chi fa la notte non puo avere altri turni lo stesso giorno'),
    ('smonto notte', 'incompatibilita', 'notte', '*', 1, 'SI',
     'Il giorno dopo la notte e di riposo obbligatorio'),
    ('doppio turno', 'incompatibilita', '*', '*', 0, 'SI',
     'Una persona non puo coprire due turni lo stesso giorno'),
    ('desiderata disatteso', 'segnalazione', '*', '*', 0, 'NO',
     'Colora la cella per farlo notare, non impedisce l inserimento'),
)


def scrivi_definizioni(wb, turni):
    """
    Le tabelle che definiscono il vocabolario: sedi, metodiche, fasce.

    Stanno per prime nella cartella di lavoro perche' sono cio' a cui tutto
    il resto si riferisce: un turno cita una sede e una metodica, una regola
    cita una fascia.

    Args:
        wb: cartella di lavoro di destinazione.
        turni (list): i turni, da cui si ricavano le metodiche in uso.
    """
    scrivi_tabella(wb, 'Presidi', 'T_Presidi',
                   ['codice', 'nome', 'descrizione'],
                   [list(presidio) for presidio in PRESIDI])

    ws = scrivi_tabella(
        wb, 'Sedi', 'T_Sedi',
        ['codice', 'nome', 'presidio', 'tipologia', 'conta_nei_dovuti',
         'pagamento_separato', 'descrizione'],
        [list(sede) for sede in SEDI])
    aggiungi_tendina(ws, ['C'], '=elenco_presidi', 2, len(SEDI) + 1)
    aggiungi_tendina(ws, ['D'], f'"{",".join(TIPOLOGIE_SEDE)}"',
                     2, len(SEDI) + 1)
    aggiungi_tendina(ws, ['E', 'F'], f'"{SI_NO}"', 2, len(SEDI) + 1)

    scrivi_tabella(wb, 'Metodiche', 'T_Metodiche',
                   ['metodica', 'descrizione'],
                   [[metodica, DESCRIZIONI_METODICHE.get(metodica, '')]
                    for metodica in sorted({t['metodica'] for t in turni
                                            if t['metodica']})])

    scrivi_tabella(wb, 'Fasce', 'T_Fasce',
                   ['fascia', 'concetto', 'inizio', 'fine', 'pausa_minuti',
                    'peso', 'solo_su_richiesta'],
                   [list(fascia) for fascia in FASCE])


def scrivi_turni(wb, turni):
    """La tabella dei turni, con le tendine sulle colonne vincolate."""
    campi = ('id', 'etichetta', 'nome_visualizzato', 'postazione',
             'sede', 'metodica',
             'fascia', 'sigla', 'riempimento', 'necessita', 'giorni',
             'festivi', 'superfestivi', 'ordine', 'attivo')

    ws = scrivi_tabella(wb, 'Turni', 'T_Turni', list(campi),
                        [[turno[campo] for campo in campi]
                         for turno in turni])

    ultima = len(turni) + 1
    aggiungi_tendina(ws, ['E'], '=elenco_sedi', 2, ultima)
    aggiungi_tendina(ws, ['F'], '=elenco_metodiche', 2, ultima)
    aggiungi_tendina(ws, ['G'], '=elenco_fasce', 2, ultima)
    aggiungi_tendina(ws, ['I'], '"obbligatorio,opzionale,chiuso"', 2, ultima)
    aggiungi_tendina(ws, ['L', 'M', 'O'], f'"{SI_NO}"', 2, ultima)

    return ws


def scrivi_persone(wb, persone):
    """La tabella delle persone: appartenenza e vincolo di sede."""
    ws = scrivi_tabella(
        wb, 'Persone', 'T_Persone',
        ['acronimo', 'cognome', 'nome', 'presidio',
         'solo_presidio_proprio', 'attivo', 'note'],
        [[p['acronimo'], p['cognome'], p['nome'], '', 'NO', 'SI', '']
         for p in persone])

    aggiungi_tendina(ws, ['D'], '=elenco_presidi', 2, len(persone) + 1)
    aggiungi_tendina(ws, ['E', 'F'], f'"{SI_NO}"', 2, len(persone) + 1)

    return ws


def scrivi_tetti(wb):
    """
    I tetti mensili: quante volte al massimo una persona fa una certa cosa.

    Un tetto vale su qualsiasi bersaglio, non solo sulle aggiuntive: «al
    massimo 4 notti a Tizio» e «al massimo 4 aggiuntive a testa» sono lo
    stesso vincolo scritto su due bersagli diversi. Lasciare `persona` in
    bianco significa "vale per tutti", cosi' il tetto generale e la deroga
    per la singola persona stanno nella stessa tabella.

    Args:
        wb: cartella di lavoro di destinazione.

    Returns:
        Il foglio creato.
    """
    # La prima riga mostra il meccanismo: tetto generale sulle aggiuntive,
    # per chiunque. Il numero lo decide chi compila.
    righe = [['', f'tipologia{SEPARATORE_LIVELLO}aggiuntiva', '',
              'Tetto generale: vale per tutti']]
    righe += [['', '', '', ''] for _ in range(RIGHE_TETTI_LIBERE)]

    ws = scrivi_tabella(wb, 'Tetti', 'T_Tetti',
                        ['persona', 'bersaglio', 'max_mese', 'note'], righe)

    ultima = len(righe) + 1
    aggiungi_tendina(ws, ['A'], '=elenco_persone', 2, ultima)
    aggiungi_tendina(ws, ['B'], '=elenco_bersagli', 2, ultima)

    return ws


def scrivi_riferimenti(wb, bersagli, richieste):
    """Bersagli, tipi di richiesta e regole assolute."""
    scrivi_tabella(wb, 'Bersagli', 'T_Bersagli',
                   ['bersaglio', 'livello', 'nome', 'descrizione'], bersagli)

    scrivi_tabella(wb, 'Richieste', 'T_Richieste',
                   ['sigla', 'tipo', 'fascia', 'conta_ore',
                    'anche_festivi', 'note'], richieste)

    scrivi_tabella(wb, 'Regole', 'T_Regole',
                   ['nome', 'tipo', 'fascia_a', 'fascia_b', 'offset_giorni',
                    'blocca', 'note'], [list(regola) for regola in REGOLE])


def scrivi_parametri(wb, mese, anno):
    """I parametri globali, compresi i pesi delle preferenze morbide."""
    parametri = (
        ('mese', mese, 'Mese del calendario in lavorazione'),
        ('anno', anno, 'Anno del calendario in lavorazione'),
        ('peso_evita', PESO_EVITA,
         'Quanto penalizza un "evita", in centesimi di turno'),
        ('peso_preferisci', PESO_PREFERISCI,
         'Quanto premia un "preferisci", in centesimi di turno'),
        ('max_giorni_consecutivi', 6, 'Giorni lavorativi di fila al massimo'),
        ('max_turni_giorno', 1, 'Turni per persona al giorno'),
        ('max_festivi_mese', 4, 'Turni in giorni festivi al mese'),
        ('giorno_chiuso', 'domenica',
         'Il giorno non lavorativo settimanale, come da modello originale'),
        ('turno_tipo_minuti', 380,
         'Unita di misura del peso: 6h20 di lavoro effettivo'),
        ('pausa_default_minuti', 10, 'Pausa obbligatoria, si somma alla netta'),
    )

    scrivi_tabella(wb, 'Parametri', 'T_Parametri',
                   ['chiave', 'valore', 'descrizione'],
                   [list(parametro) for parametro in parametri])


def _note_di_lettura(turni, persone, postazioni):
    """Il testo del foglio Leggimi: cosa e' letto e cosa e' proposto."""
    proposti = 'riempimento, necessita, giorni, festivi, superfestivi'
    return (
        ('Da dove viene', f'Generato da {MODELLO_ORIGINE}, foglio Tabelle, '
         f'con genera_config_turni.py. L originale non e stato modificato.'),
        ('Letto dal foglio', f'{len(turni)} turni raggruppati in '
         f'{postazioni} postazioni, {len(persone)} persone, sigle, '
         f'metodiche e tipi di richiesta.'),
        ('Proposto, da rivedere', f'Le colonne {proposti} di T_Turni: il '
         f'foglio originale non contiene questa informazione. I valori '
         f'attuali sono un default ragionevole, non un dato letto.'),
        ('Da compilare', 'T_Persone: presidio e solo_presidio_proprio. '
         'T_Preferenze: tutte le regole.'),
        ('Come si scrive un bersaglio', 'Sempre "livello: nome", per esempio '
         '"sede: S.G." oppure "turno: ADD. TC · mattina". Il pezzo prima dei '
         'due punti dice il livello, quello dopo il nome. Livello e nome '
         'stanno comunque gia separati nelle colonne di T_Bersagli. '
         'I cinque livelli sono '
         'fascia, tipologia, sede, metodica, turno. Un turno come '
         'bersaglio e una fascia sola: S.G. DEA 1 mattina e S.G. DEA 1 '
         'pomeriggio si escludono separatamente.'),
        ('Ordine della tendina', 'Dal generale al particolare: fasce, tipologie, '
         'sedi, metodiche, turni.'),
        ('Sedi aggiuntive', 'I turni di una sede con tipologia aggiuntiva '
         'non concorrono al monte turni dovuti e si retribuiscono a parte '
         '(colonne conta_nei_dovuti e pagamento_separato di T_Sedi). Per '
         'sapere quante aggiuntive ha fatto una persona si contano i suoi '
         'turni la cui sede ha tipologia aggiuntiva.'),
        ('Tetti mensili', 'T_Tetti limita quante volte al mese una persona '
         'puo fare un dato bersaglio. Persona in bianco significa "vale per '
         'tutti"; una riga con la persona scritta deroga al tetto generale.'),
        ('Modi delle regole', 'mai e solo sono assoluti; evita e preferisci '
         'pesano quanto dice T_Parametri. Piu regole solo su assi diversi '
         'si intersecano. mai batte solo batte i modi morbidi.'),
    )


def main():
    """Legge il modello e scrive la cartella di lavoro di configurazione."""
    origine = openpyxl.load_workbook(MODELLO_ORIGINE, data_only=True)
    tabelle = origine[FOGLIO_TABELLE]

    turni = completa_turni(unifica_postazioni(leggi_turni(tabelle)))
    persone = leggi_persone(tabelle)
    bersagli, doppioni = costruisci_bersagli(turni)
    richieste = costruisci_richieste(tabelle)
    mese = origine['Inserimento']['Y15'].value
    anno = origine['Inserimento']['AA15'].value

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # L'ordine delle chiamate e' l'ordine dei fogli: prima cio' che
    # definisce il vocabolario, poi cio' che lo usa.
    scrivi_definizioni(wb, turni)
    scrivi_turni(wb, turni)
    scrivi_persone(wb, persone)
    scrivi_preferenze(wb, persone)
    scrivi_tetti(wb)
    scrivi_riferimenti(wb, bersagli, richieste)
    scrivi_parametri(wb, mese, anno)

    festivita = calcola_festivita(anno)
    _, celle_aperte = scrivi_calendario(wb, turni, mese, anno, festivita,
                                        len(persone), len(richieste))
    scrivi_desiderata(wb, persone, mese, anno, festivita)

    for nome, foglio in (('elenco_bersagli', 'Bersagli'),
                         ('elenco_persone', 'Persone'),
                         ('elenco_sedi', 'Sedi'),
                         ('elenco_presidi', 'Presidi'),
                         ('elenco_metodiche', 'Metodiche'),
                         ('elenco_fasce', 'Fasce'),
                         ('elenco_richieste', 'Richieste')):
        registra_nome(wb, nome, _intervallo_dinamico(foglio))

    postazioni = len({turno['postazione'] for turno in turni})
    scrivi_leggimi(wb, _note_di_lettura(turni, persone, postazioni))
    wb.save(FILE_USCITA)

    print(f'Scritto {FILE_USCITA}')
    print(f'  fogli: {", ".join(wb.sheetnames)}')
    print(f'  turni {len(turni)} · postazioni {postazioni} · '
          f'persone {len(persone)} · bersagli {len(bersagli)}')
    print(f'  doppioni nei bersagli: {doppioni or "nessuno"}')
    print(f'  griglia: {celle_aperte} celle da coprire in {mese:02d}/{anno}')



# Le festivita' fisse, come le calcola il foglio `festivi` dell'originale:
# (giorno, mese, nome). San Pietro e Paolo e' il patrono di Roma.
FESTIVITA_FISSE = (
    (1,  1,  'Capodanno'),      (6,  1,  'Epifania'),
    (25, 4,  'Liberazione'),    (1,  5,  'Festa del lavoro'),
    (2,  6,  'Repubblica'),     (29, 6,  'San Pietro e Paolo'),
    (15, 8,  'Ferragosto'),     (1,  11, 'Ognissanti'),
    (8,  12, 'Immacolata'),     (25, 12, 'Natale'),
    (26, 12, 'Santo Stefano'),
)

# Le festivita' che si contano dalla Pasqua: (offset in giorni, nome).
FESTIVITA_MOBILI = ((0, 'Pasqua'), (1, 'Pasquetta'))

# Le iniziali dei giorni, nell'ordine di date.weekday(): lunedi' = 0.
INIZIALI_GIORNI = 'LMMGVSD'
DOMENICA = 6

# Come si colorano le celle chiuse e le intestazioni della griglia.
GRIGIO_CHIUSO = 'FFD9D9D9'
GIALLO_FESTIVO = 'FFFFF2CC'
GRIGIO_INTESTAZIONE = 'FFF2F2F2'

MESI = ('gennaio', 'febbraio', 'marzo', 'aprile', 'maggio', 'giugno',
        'luglio', 'agosto', 'settembre', 'ottobre', 'novembre', 'dicembre')

# Le richieste che nominano una fascia: chi chiede M e finisce di
# pomeriggio ha un desiderata disatteso, e la griglia deve dirlo.
FASCIA_RICHIESTA = {'M': 'mattina', 'P': 'pomeriggio',
                    'N': 'notte', 'L': 'lunga'}

# I colori delle segnalazioni. I gravi rompono una regola assoluta, gli
# avvisi disattendono una richiesta: due gravita', due letture a colpo
# d'occhio.
COLORE_GRAVE_SFONDO = 'FFFFC7CE'
COLORE_GRAVE_TESTO = 'FF9C0006'
COLORE_ASSENZA_SFONDO = 'FFFFEB9C'
COLORE_ASSENZA_TESTO = 'FF9C6500'
COLORE_FASCIA_SFONDO = 'FFE4DFEC'
COLORE_FASCIA_TESTO = 'FF7030A0'

# Le colonne di servizio del Calendario: portano la fascia e l'id di ogni
# riga, cosi' le regole sanno che turno stanno guardando senza bisogno di
# un foglio parallelo. Restano nascoste.
COLONNA_FASCIA_RIGA = 34
COLONNA_ID_RIGA = 35


# Le sezioni della griglia, nell'ordine e con i colori del modello
# originale: (titolo, tipologia della sede, fascia, colore del titolo,
# colore della colonna dei nomi).
SEZIONI = (
    ('MATTINA',               'ordinario',  'mattina',    'FFFFC000',
     'FF4472C4'),
    ('POMERIGGIO',            'ordinario',  'pomeriggio', 'FF70AD47',
     'FF4472C4'),
    ('NOTTE',                 'ordinario',  'notte',      'FF4472C4',
     'FF5B9BD5'),
    ('AGGIUNTIVA MATTINA',    'aggiuntiva', 'mattina',    'FFFFC000',
     'FFFFC000'),
    ('AGGIUNTIVA POMERIGGIO', 'aggiuntiva', 'pomeriggio', 'FF4472C4',
     'FFFEE2E8'),
)

# Il titolo di sezione si ripete tre volte in orizzontale, come
# nell'originale: cosi' resta leggibile anche scorrendo a meta' mese.
BANDE_TITOLO = 3

# Misure prese dal modello: colonne, altezze, corpi dei caratteri.
LARGHEZZA_SIGLA = 12
LARGHEZZA_NOME = 19.7
LARGHEZZA_GIORNO = 8.7
ALTEZZA_TITOLO = 15.75
ALTEZZA_RIGA = 13.8
ALTEZZA_SPAZIATORE = 5.25

# Le prime tre righe sono testata e spaziatore: i dati partono dopo.
PRIMA_RIGA_DATI = 4
CORPO_GRIGLIA = 8
CORPO_NOME = 10
CORPO_TITOLO = 12
CORPO_INTESTAZIONE = 9
FORMATO_MESE = 'mmmm yyyy'
FORMATO_GIORNO = 'dd'



def calcola_pasqua(anno):
    """
    La domenica di Pasqua, con l'algoritmo gregoriano anonimo.

    E' la forma canonica (Meeus/Jones/Butcher): le lettere sono quelle
    della pubblicazione originale e non hanno un significato che convenga
    ribattezzare. Verificata su 2024-03-31, 2025-04-20, 2026-04-05.

    Args:
        anno (int): l'anno.

    Returns:
        datetime.date: la data di Pasqua.
    """
    a = anno % 19
    b, c = divmod(anno, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451

    mese, giorno = divmod(h + l - 7 * m + 114, 31)

    return datetime.date(anno, mese, giorno + 1)


def calcola_festivita(anno):
    """
    Le date festive di un anno: le fisse piu' quelle legate alla Pasqua.

    Le domeniche non stanno qui — quelle le sa il calendario.

    Args:
        anno (int): l'anno.

    Returns:
        dict: data -> nome della festivita'.
    """
    festivita = {datetime.date(anno, mese, giorno): nome
                 for giorno, mese, nome in FESTIVITA_FISSE}

    pasqua = calcola_pasqua(anno)
    for scarto, nome in FESTIVITA_MOBILI:
        festivita[pasqua + datetime.timedelta(days=scarto)] = nome

    return festivita


def _giorni_del_mese(mese, anno):
    """Le date del mese, in ordine."""
    ultimo = calendar.monthrange(anno, mese)[1]

    return [datetime.date(anno, mese, giorno)
            for giorno in range(1, ultimo + 1)]


def turno_aperto(turno, data, festivita):
    """
    Se un turno va coperto in un dato giorno.

    Tre condizioni in and: il turno e' attivo, la maschera dei giorni
    ammette quel giorno della settimana, e — se e' domenica o festivita' —
    il turno dichiara di aprire anche in quei giorni.

    Args:
        turno (dict): la riga del turno.
        data (datetime.date): il giorno.
        festivita (dict): le date festive dell'anno.

    Returns:
        bool: True se la cella va riempita.
    """
    if turno['attivo'] != 'SI' or turno['riempimento'] == 'chiuso':
        return False

    if turno['giorni'][data.weekday()] == '-':
        return False

    if data.weekday() == DOMENICA and turno['festivi'] != 'SI':
        return False

    if data in festivita and turno['superfestivi'] != 'SI':
        return False

    return True


def _bordo_sottile():
    """Il riquadro sottile che il modello usa su tutte le celle."""
    lato = openpyxl.styles.Side(style='thin')

    return openpyxl.styles.Border(left=lato, right=lato,
                                  top=lato, bottom=lato)


def _intesta_calendario(ws, giorni, festivita, prima_colonna, etichetta):
    """
    Le due righe di testata: le date e i giorni della settimana.

    Ricalca il modello: il mese scritto per esteso a sinistra, le date in
    formato `dd`, le iniziali dei giorni sotto, tutto incorniciato e
    centrato. Le colonne di domenica e festivita' si colorano, come li'.

    Args:
        ws: il foglio.
        giorni (list): le date del mese.
        festivita (dict): le date festive dell'anno.
        prima_colonna (int): colonna della prima data.
        etichetta (str): cosa scrivere nella prima colonna.
    """
    bordo = _bordo_sottile()
    centrato = openpyxl.styles.Alignment(horizontal='center')
    giallo = openpyxl.styles.PatternFill('solid', fgColor=GIALLO_FESTIVO)

    intestazione = ws.cell(1, prima_colonna - 1, giorni[0])
    intestazione.number_format = FORMATO_MESE
    intestazione.font = openpyxl.styles.Font(bold=True,
                                             size=CORPO_INTESTAZIONE)
    intestazione.alignment = centrato
    intestazione.border = bordo

    sotto = ws.cell(2, prima_colonna - 1, etichetta)
    sotto.font = openpyxl.styles.Font(bold=True, size=CORPO_INTESTAZIONE)
    sotto.alignment = centrato
    sotto.border = bordo

    for scarto, data in enumerate(giorni):
        colonna = prima_colonna + scarto
        festivo = data.weekday() == DOMENICA or data in festivita

        for riga, valore, formato in (
                (1, data, FORMATO_GIORNO),
                (2, INIZIALI_GIORNI[data.weekday()], 'General')):
            cella = ws.cell(riga, colonna, valore)
            cella.number_format = formato
            cella.font = openpyxl.styles.Font(size=CORPO_GRIGLIA)
            cella.alignment = centrato
            cella.border = bordo
            if festivo:
                cella.fill = giallo

        ws.column_dimensions[get_column_letter(colonna)].width = \
            LARGHEZZA_GIORNO

    ws.row_dimensions[3].height = ALTEZZA_SPAZIATORE


def _scrivi_titolo_sezione(ws, riga, titolo, colore, prima_colonna, giorni):
    """
    La banda che apre una sezione, ripetuta in orizzontale come nel modello.

    Args:
        ws: il foglio.
        riga (int): la riga della banda.
        titolo (str): il testo, es. `MATTINA`.
        colore (str): colore di riempimento in formato aRGB.
        prima_colonna (int): colonna della prima data.
        giorni (int): quante colonne di date ci sono.
    """
    larghezza = giorni // BANDE_TITOLO
    riempimento = openpyxl.styles.PatternFill('solid', fgColor=colore)
    ws.row_dimensions[riga].height = ALTEZZA_TITOLO

    for banda in range(BANDE_TITOLO):
        da = prima_colonna - 1 + banda * larghezza
        # L'ultima banda si prende le colonne avanzate dalla divisione.
        a = (prima_colonna - 1 + giorni if banda == BANDE_TITOLO - 1
             else da + larghezza - 1)

        cella = ws.cell(riga, da, titolo)
        cella.fill = riempimento
        cella.font = openpyxl.styles.Font(bold=True, size=CORPO_TITOLO)
        cella.alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells(start_row=riga, start_column=da,
                       end_row=riga, end_column=a)


def _tipologia_per_sede():
    """Da codice di sede a tipologia, per sapere se un turno e' aggiuntivo."""
    return {sede[0]: sede[3] for sede in SEDI}


def scrivi_calendario(wb, turni, mese, anno, festivita, numero_persone,
                      numero_richieste):
    """
    La griglia dei turni, nella veste del modello originale.

    Le righe non sono piu' cablate — "mattina dalla 22 alla 39" — ma
    generate da T_Turni, raggruppate nelle stesse sezioni del modello e
    ordinate per la colonna `ordine`. Aggiungere un turno significa
    aggiungere una riga alla tabella.

    Le celle grigie sono chiuse: quel turno, quel giorno, non si copre.

    Args:
        wb: cartella di lavoro di destinazione.
        turni (list): i turni completati.
        mese (int), anno (int): il mese da programmare.
        festivita (dict): le date festive dell'anno.
        numero_persone (int): quante righe ha il foglio Desiderata.
        numero_richieste (int): quante righe ha T_Richieste.

    Returns:
        tuple: (foglio, numero di celle da coprire).
    """
    ws = wb.create_sheet('Calendario')
    giorni = _giorni_del_mese(mese, anno)
    prima_colonna = 3

    _intesta_calendario(ws, giorni, festivita, prima_colonna,
                        'Giorni Settimana')
    ws.column_dimensions['A'].width = LARGHEZZA_SIGLA
    ws.column_dimensions['B'].width = LARGHEZZA_NOME

    tipologie = _tipologia_per_sede()
    riga = PRIMA_RIGA_DATI
    aperte = 0

    for titolo, tipologia, fascia, colore_titolo, colore_nome in SEZIONI:
        della_sezione = [t for t in turni
                         if t['attivo'] == 'SI' and t['fascia'] == fascia
                         and tipologie.get(t['sede']) == tipologia]
        if not della_sezione:
            continue

        _scrivi_titolo_sezione(ws, riga, titolo, colore_titolo,
                               prima_colonna, len(giorni))
        riga += 1
        aperte += _scrivi_righe_turno(ws, della_sezione, riga, giorni,
                                      festivita, prima_colonna, colore_nome)
        riga += len(della_sezione)

    ultima_riga = riga - 1
    ultimo_giorno = get_column_letter(prima_colonna + len(giorni) - 1)

    ws.freeze_panes = (f'{get_column_letter(prima_colonna)}'
                       f'{PRIMA_RIGA_DATI}')
    aggiungi_tendina(ws, [get_column_letter(prima_colonna + scarto)
                          for scarto in range(len(giorni))],
                     'elenco_persone', PRIMA_RIGA_DATI, ultima_riga)

    for colonna in (COLONNA_FASCIA_RIGA, COLONNA_ID_RIGA):
        ws.column_dimensions[get_column_letter(colonna)].hidden = True

    _applica_segnalazioni(ws, ultima_riga, ultimo_giorno,
                          numero_persone, numero_richieste)
    _scrivi_legenda(ws, ultima_riga + 2)

    return ws, aperte


def _applica_segnalazioni(ws, ultima_riga, ultimo_giorno,
                          numero_persone, numero_richieste):
    """
    Attacca alla griglia le regole che colorano gli errori.

    Args:
        ws: il foglio Calendario.
        ultima_riga (int): ultima riga della griglia.
        ultimo_giorno (str): lettera dell'ultima colonna dei giorni.
        numero_persone (int): quante righe ha il foglio Desiderata.
        numero_richieste (int): quante righe ha la tabella delle richieste.
    """
    richiesta = _richiesta_del_giorno(
        PRIMA_RIGA_DATI, PRIMA_RIGA_DATI - 1 + numero_persone, ultimo_giorno)

    for prima_cella, formula, stile, _ in _regole_segnalazione(
            ultima_riga, richiesta, numero_richieste + 1):
        colonna = prima_cella[0]
        intervallo = (f'{colonna}{PRIMA_RIGA_DATI}:'
                      f'{ultimo_giorno}{ultima_riga}')
        regola = openpyxl.formatting.rule.Rule(
            type='expression', dxf=stile, formula=[formula])
        ws.conditional_formatting.add(intervallo, regola)


def _scrivi_legenda(ws, riga):
    """La legenda dei colori, sotto la griglia."""
    ws.cell(riga, 2, 'LEGENDA SEGNALAZIONI').font = \
        openpyxl.styles.Font(bold=True, size=CORPO_NOME)

    voci = ((COLORE_GRAVE_SFONDO, COLORE_GRAVE_TESTO,
             'Regola assoluta violata: doppio turno, notte con altro '
             'turno, smonto notte non rispettato'),
            (COLORE_ASSENZA_SFONDO, COLORE_ASSENZA_TESTO,
             'Assegnato in un giorno per cui aveva chiesto un assenza'),
            (COLORE_FASCIA_SFONDO, COLORE_FASCIA_TESTO,
             'Fascia diversa da quella richiesta'),
            (GRIGIO_CHIUSO, None, 'Cella chiusa: quel turno quel giorno '
             'non si copre'))

    for scarto, (sfondo, colore, descrizione) in enumerate(voci):
        campione = ws.cell(riga + 1 + scarto, 2, 'esempio')
        campione.fill = openpyxl.styles.PatternFill('solid', fgColor=sfondo)
        campione.alignment = openpyxl.styles.Alignment(horizontal='center')
        if colore:
            campione.font = openpyxl.styles.Font(bold=True, color=colore)

        ws.cell(riga + 1 + scarto, 3, descrizione).font = \
            openpyxl.styles.Font(size=CORPO_GRIGLIA)


def _scrivi_righe_turno(ws, turni, prima_riga, giorni, festivita,
                        prima_colonna, colore_nome):
    """
    Le righe di una sezione: sigla, nome del turno e celle del mese.

    Args:
        ws: il foglio.
        turni (list): i turni della sezione, in ordine.
        prima_riga (int): riga da cui partire.
        giorni (list): le date del mese.
        festivita (dict): le date festive.
        prima_colonna (int): colonna della prima data.
        colore_nome (str): riempimento della colonna dei nomi.

    Returns:
        int: quante celle sono aperte in queste righe.
    """
    bordo = _bordo_sottile()
    centrato = openpyxl.styles.Alignment(horizontal='center')
    grigio = openpyxl.styles.PatternFill('solid', fgColor=GRIGIO_CHIUSO)
    riempimento = openpyxl.styles.PatternFill('solid', fgColor=colore_nome)
    aperte = 0

    for scarto, turno in enumerate(turni):
        riga = prima_riga + scarto
        ws.row_dimensions[riga].height = ALTEZZA_RIGA

        sigla = ws.cell(riga, 1, turno['sigla'])
        sigla.font = openpyxl.styles.Font(size=CORPO_GRIGLIA)
        sigla.alignment = centrato

        # Le colonne di servizio: dicono alle regole che turno e'
        # questa riga, senza che debbano contare le righe.
        ws.cell(riga, COLONNA_FASCIA_RIGA, turno['fascia'])
        ws.cell(riga, COLONNA_ID_RIGA, turno['id'])

        nome = ws.cell(riga, 2, turno['nome_visualizzato'])
        nome.fill = riempimento
        nome.font = openpyxl.styles.Font(bold=True, size=CORPO_NOME)
        nome.alignment = centrato
        nome.border = bordo

        for indice, data in enumerate(giorni):
            cella = ws.cell(riga, prima_colonna + indice)
            cella.font = openpyxl.styles.Font(size=CORPO_GRIGLIA)
            cella.alignment = centrato
            cella.border = bordo
            if turno_aperto(turno, data, festivita):
                aperte += 1
            else:
                cella.fill = grigio

    return aperte


def _stile_segnalazione(sfondo, testo):
    """Lo stile differenziale che una regola applica quando scatta."""
    return openpyxl.styles.differential.DifferentialStyle(
        font=openpyxl.styles.Font(bold=True, color=testo),
        fill=openpyxl.styles.PatternFill(bgColor=sfondo))


def _richiesta_del_giorno(riga_desiderata, ultima_persona, ultimo_giorno):
    """
    La formula che pesca cosa aveva chiesto, quel giorno, chi sta in cella.

    Le due griglie hanno i giorni nelle stesse colonne, quindi lo stesso
    scarto vale per entrambe e non serve cercare la data.

    Args:
        riga_desiderata (int): prima riga di dati del foglio Desiderata.
        ultima_persona (int): ultima riga di dati del foglio Desiderata.
        ultimo_giorno (str): lettera dell'ultima colonna dei giorni.

    Returns:
        str: il frammento di formula, senza uguale.
    """
    return (f'IFERROR(INDEX(Desiderata!$C${riga_desiderata}:'
            f'${ultimo_giorno}${ultima_persona},'
            f'MATCH(C{PRIMA_RIGA_DATI},Desiderata!$B${riga_desiderata}:'
            f'$B${ultima_persona},0),COLUMN()-2),"")')


def _regole_segnalazione(ultima_riga, richiesta, ultima_richiesta):
    """
    Le regole che colorano gli errori, in ordine di gravita'.

    Sono le stesse che l'originale otteneva con il foglio Matrici e
    trentacinque formattazioni condizionali, ma scritte in modo che non
    dipendano dalla posizione delle righe: dove sta la notte lo dice la
    colonna di servizio, non il numero di riga.

    Ogni intervallo e' delimitato. Un riferimento a colonna intera qui si
    paga caro: la formula gira su millecinquecento celle, e scandire un
    milione di righe per ognuna rende il foglio inusabile.

    Args:
        ultima_riga (int): ultima riga della griglia.
        richiesta (str): formula che ricava il desiderata del giorno.
        ultima_richiesta (int): ultima riga della tabella dei tipi di
            richiesta.

    Returns:
        list: tuple (prima cella, formula, stile, descrizione).
    """
    sigle = f'Richieste!$A$2:$A${ultima_richiesta}'
    tipi = f'Richieste!$B$2:$B${ultima_richiesta}'
    fasce_richieste = f'Richieste!$C$2:$C${ultima_richiesta}'

    fasce = (f'${get_column_letter(COLONNA_FASCIA_RIGA)}$'
             f'{PRIMA_RIGA_DATI}:${get_column_letter(COLONNA_FASCIA_RIGA)}$'
             f'{ultima_riga}')
    fascia_riga = (f'${get_column_letter(COLONNA_FASCIA_RIGA)}'
                   f'{PRIMA_RIGA_DATI}')
    colonna = f'C${PRIMA_RIGA_DATI}:C${ultima_riga}'
    # Ancorata alla colonna D, quindi il giorno prima e' C: Excel trasla
    # il riferimento insieme alla cella che valuta.
    precedente = f'C${PRIMA_RIGA_DATI}:C${ultima_riga}'
    cella = f'C{PRIMA_RIGA_DATI}'
    grave = _stile_segnalazione(COLORE_GRAVE_SFONDO, COLORE_GRAVE_TESTO)

    # Il tipo e la fascia della richiesta si prendono con una ricerca
    # sola, invece di due conteggi: e' la stessa risposta a meta' costo.
    tipo_chiesto = f'IFERROR(INDEX({tipi},MATCH({richiesta},{sigle},0)),"")'
    fascia_chiesta = (f'IFERROR(INDEX({fasce_richieste},'
                      f'MATCH({richiesta},{sigle},0)),"")')

    return [
        (f'C{PRIMA_RIGA_DATI}',
         f'AND({cella}<>"",COUNTIF({colonna},{cella})>1)',
         grave, 'La stessa persona due volte nello stesso giorno'),

        (f'C{PRIMA_RIGA_DATI}',
         f'AND({cella}<>"",'
         f'COUNTIFS({fasce},"notte",{colonna},{cella})>0,'
         f'COUNTIFS({fasce},"<>notte",{colonna},{cella})>0)',
         grave, 'Notte e altro turno lo stesso giorno'),

        (f'D{PRIMA_RIGA_DATI}',
         f'AND(D{PRIMA_RIGA_DATI}<>"",'
         f'COUNTIFS({fasce},"notte",{precedente},D{PRIMA_RIGA_DATI})>0)',
         grave, 'Smonto notte: ha lavorato la notte prima'),

        (f'C{PRIMA_RIGA_DATI}',
         f'AND({cella}<>"",{tipo_chiesto}="assenza")',
         _stile_segnalazione(COLORE_ASSENZA_SFONDO, COLORE_ASSENZA_TESTO),
         'Assegnato in un giorno di assenza richiesta'),

        (f'C{PRIMA_RIGA_DATI}',
         f'AND({cella}<>"",{fascia_riga}<>"",{fascia_chiesta}<>"",'
         f'{fascia_chiesta}<>{fascia_riga})',
         _stile_segnalazione(COLORE_FASCIA_SFONDO, COLORE_FASCIA_TESTO),
         'Fascia diversa da quella richiesta'),
    ]


def scrivi_desiderata(wb, persone, mese, anno, festivita):
    """
    Le richieste dei lavoratori: una riga per persona, una per giorno.

    E' l'ingresso del solver insieme alle preferenze: qui sta cosa la
    persona ha chiesto per quel giorno preciso, mentre T_Preferenze tiene
    cio' che vale sempre. La disposizione ricalca il blocco desiderata del
    modello: acronimi in colonna B, giorni da C in poi.

    Args:
        wb: cartella di lavoro di destinazione.
        persone (list): le persone.
        mese (int), anno (int): il mese da programmare.
        festivita (dict): le date festive dell'anno.

    Returns:
        Il foglio creato.
    """
    ws = wb.create_sheet('Desiderata')
    giorni = _giorni_del_mese(mese, anno)
    prima_colonna = 3

    _intesta_calendario(ws, giorni, festivita, prima_colonna, 'MEDICO')
    ws.column_dimensions['A'].width = LARGHEZZA_SIGLA
    ws.column_dimensions['B'].width = LARGHEZZA_NOME

    bordo = _bordo_sottile()
    centrato = openpyxl.styles.Alignment(horizontal='center')
    grigio = openpyxl.styles.PatternFill('solid', fgColor=GRIGIO_INTESTAZIONE)

    for scarto, persona in enumerate(persone):
        riga = PRIMA_RIGA_DATI + scarto
        ws.row_dimensions[riga].height = ALTEZZA_RIGA

        cella = ws.cell(riga, 2, persona['acronimo'])
        cella.fill = grigio
        cella.font = openpyxl.styles.Font(bold=True, size=CORPO_NOME)
        cella.alignment = centrato
        cella.border = bordo

        for indice in range(len(giorni)):
            giorno = ws.cell(riga, prima_colonna + indice)
            giorno.font = openpyxl.styles.Font(size=CORPO_GRIGLIA)
            giorno.alignment = centrato
            giorno.border = bordo

    ultima = PRIMA_RIGA_DATI - 1 + len(persone)
    ws.freeze_panes = (f'{get_column_letter(prima_colonna)}'
                       f'{PRIMA_RIGA_DATI}')
    aggiungi_tendina(ws, [get_column_letter(prima_colonna + scarto)
                          for scarto in range(len(giorni))],
                     'elenco_richieste',
                     PRIMA_RIGA_DATI, ultima)

    return ws


if __name__ == '__main__':
    main()
