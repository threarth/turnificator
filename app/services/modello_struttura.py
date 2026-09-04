"""
app/services/modello_struttura.py — la struttura turni letta da un foglio Excel.

Nuova feature. Chi arriva da un foglio di calcolo ha gia' tutto scritto li':
i turni, come sono raggruppati, chi ci lavora, con che tipologie. Ridigitarlo
nella configurazione guidata e' un lavoro lungo e un'occasione di sbagliare.

Il foglio `Tabelle` e' la struttura in forma leggibile dalla macchina, e non
per caso: le righe della griglia in `Inserimento` sono generate da li' con
formule matriciali. Importare da `Tabelle` significa quindi ottenere una
struttura che combacia riga per riga con il foglio, ed e' cio' che permette
poi di riesportare il mese dentro lo stesso modello.

Cosa si legge:

    colonne A/B/C   turni della mattina: nome, tipologia, sigla di gruppo
    colonne E/F/G   turni del pomeriggio
    colonne I/J/K   turni della notte
    righe 27+       gli aggiuntivi, mattina e pomeriggio
    colonne N/O/P   le persone: acronimo, cognome, nome
    colonne R/S     i tipi di richiesta: desiderata e assenze

La **struttura** e' il luogo, e lo dice la prima parola del nome del turno:
«S.G. DEA 1», «ADD. TC», «S.M. MX/ECO 1». Non e' la colonna del Riepilogo,
che incrocia metodica e sede ed e' un'altra cosa.

Ai turni che il luogo non lo nominano — la notte si chiama «NOTTE - 1° med.»
— la struttura arriva dagli altri turni dello stesso **posto**: `deaM;`,
`deaP;` e `deaN` sono lo stesso posto in tre momenti della giornata, e se due
di quei turni stanno a S.G. ci sta anche il terzo.

Resta una deduzione, e come tale si propone: chi importa vede le strutture
trovate e puo' correggerle o fonderle prima di creare.

Questo modulo **non tocca il database**: legge e descrive. Chi scrive decide
a parte, dopo aver mostrato all'utente cosa e' stato capito.
"""

import re
from collections import Counter

import openpyxl
from openpyxl.utils import column_index_from_string


# Il foglio da cui si legge la struttura.
FOGLIO_TABELLE = 'Tabelle'

# I blocchi di turni, per nome dell'intervallo definito nel foglio. Il modello
# li dichiara, e sono la stessa cosa che le formule matriciali di
# `Inserimento` usano per generare le righe della griglia: leggere di qui
# significa ottenere i turni nell'ordine esatto in cui la griglia li dispone.
#
# `ripiego` sono le posizioni del modello, per un foglio che i nomi non li ha.
BLOCCHI_TURNI = (
    {'fascia': 'mattina',    'nomi': 'turni_mattina',
     'sigle': 'sigle_turni_matt',      'ripiego': ('A', 2, 19, 'C')},
    {'fascia': 'pomeriggio', 'nomi': 'turni_pomeriggio',
     'sigle': 'sigle_turni_pomeriggio', 'ripiego': ('E', 2, 19, 'G')},
    {'fascia': 'notte',      'nomi': 'turni_notte',
     'sigle': 'sigle_turni_notte',      'ripiego': ('I', 2, 3, 'K')},
    {'fascia': 'mattina',    'nomi': 'agg_mattina',
     'sigle': 'sigle_agg_matt',         'ripiego': ('A', 27, 34, 'C')},
    {'fascia': 'pomeriggio', 'nomi': 'agg_pomeriggio',
     'sigle': 'sigle_agg_pom',          'ripiego': ('E', 27, 34, 'G')},
)

# La tipologia sta sempre nella colonna accanto al nome.
SCARTO_COLONNA_TIPOLOGIA = 1

# Le persone e i tipi di richiesta, sempre per nome dell'intervallo.
NOME_PERSONE = 'medici'
RIPIEGO_PERSONE = ('N', 2, 42)
SCARTO_COGNOME = 1
SCARTO_NOME = 2

# Tutto cio' che si puo' chiedere, e quali di quelle cose sono assenze.
NOME_RICHIESTE = 'tipo_desiderata'
RIPIEGO_RICHIESTE = ('R', 2, 13)
NOME_ASSENZE = 'tipo_assenze_romc'
RIPIEGO_ASSENZE = ('S', 2, 9)

TIPO_LAVORATIVO = 'lavorativo'
TIPO_ASSENZA = 'assenza'

# Il segnaposto con cui il foglio riempie le righe non usate.
SEGNAPOSTO = ('-', '', 'vuoto', 'chiusa', 'chiuso')


def _testo(cella):
    """Il contenuto di una cella come testo pulito, o stringa vuota."""
    if cella is None or cella.value is None:
        return ''

    return str(cella.value).strip()


def _e_segnaposto(testo):
    """La riga e' una di quelle che il foglio lascia libere."""
    return testo.lower().strip(' .;') in SEGNAPOSTO


def _intervallo(wb, nome, ripiego):
    """
    Dove sta un blocco: colonna e righe, dall'intervallo che il foglio nomina.

    Il modello dichiara i propri blocchi con un nome (`turni_mattina`,
    `medici`): leggerli di li' e' piu' solido che contare le colonne, e
    sopravvive a una riga aggiunta in mezzo. Il ripiego serve a un foglio
    costruito a mano, senza quei nomi.

    Args:
        wb: cartella di lavoro.
        nome (str): nome dell'intervallo definito.
        ripiego (tuple): (colonna, prima riga, ultima riga) del modello.

    Returns:
        tuple: (indice colonna, prima riga, ultima riga).
    """
    definito = wb.defined_names.get(nome)
    if definito is not None:
        trovato = re.search(r'\$([A-Z]+)\$(\d+):\$[A-Z]+\$(\d+)', str(definito.value))
        if trovato:
            return (column_index_from_string(trovato.group(1)),
                    int(trovato.group(2)), int(trovato.group(3)))

    colonna, prima, ultima = ripiego

    return column_index_from_string(colonna), prima, ultima


def _chiave_sigla(sigla):
    """
    La sigla privata del punto e virgola e del suffisso di fascia.

    `deaM;` e `deaP;` diventano `dea`: sono lo stesso posto, in due momenti
    della giornata. E' il ripiego per quando il Riepilogo non si legge.

    Args:
        sigla (str): sigla come sta nel foglio.

    Returns:
        str: la chiave del raggruppamento, minuscola.
    """
    pulita = sigla.strip().rstrip(';').strip()

    return re.sub(r'[mpn]$', '', pulita, flags=re.IGNORECASE).lower()


def _posto_del_nome(nome):
    """
    Il luogo scritto in testa al nome di un turno.

    «S.G. DEA 1», «ADD. TC», «S.M. MX/ECO 1»: la prima parola dice la sede, ed
    e' quella che distingue una struttura dall'altra. Un nome di una parola
    sola non nomina nessun luogo.

    Args:
        nome (str): nome del turno.

    Returns:
        tuple: (chiave confrontabile, forma da mostrare); vuote se non c'e'.
    """
    pezzi = nome.split()
    if len(pezzi) < 2:
        return '', ''

    primo = pezzi[0].strip()

    return ''.join(c for c in primo.upper() if c.isalnum()), primo


def _strutture_dei_posti(righe):
    """
    A quale struttura appartiene ciascun posto di lavoro.

    Un "posto" e' la sigla senza il suffisso di fascia: `deaM;` e `deaN` sono
    lo stesso posto in due momenti della giornata. La struttura del posto e'
    il luogo che compare piu' spesso nei nomi dei suoi turni — cosi' la notte,
    che si chiama «NOTTE - 1° med.» e la sede non la nomina, finisce dove
    stanno gli altri turni dello stesso posto.

    Args:
        righe (list): [(sigla, nome)] di tutti i turni letti.

    Returns:
        dict: chiave del posto → (chiave struttura, nome da mostrare).
    """
    per_posto = {}
    for sigla, nome in righe:
        per_posto.setdefault(_chiave_sigla(sigla), []).append(_posto_del_nome(nome))

    strutture = {}
    for posto, luoghi in per_posto.items():
        nominati = [l for l in luoghi if l[0]]
        if not nominati:
            strutture[posto] = (posto.upper(), posto.upper())
            continue

        vincitore = Counter(chiave for chiave, _ in nominati).most_common(1)[0][0]
        mostrato = next(m for c, m in nominati if c == vincitore)
        strutture[posto] = (vincitore, mostrato)

    return strutture


def _righe_turni(wb, ws):
    """
    Le righe di turno del foglio, nell'ordine in cui la griglia le dispone.

    Args:
        wb: cartella di lavoro, per gli intervalli nominati.
        ws: foglio Tabelle, a valori.

    Returns:
        tuple: (lista di righe grezze, lista di avvisi).
    """
    righe = []
    avvisi = []

    for blocco in BLOCCHI_TURNI:
        col_nome, prima, ultima = _intervallo(wb, blocco['nomi'], blocco['ripiego'][:3])
        col_sigla, _, _ = _intervallo(
            wb, blocco['sigle'], (blocco['ripiego'][3], prima, ultima)
        )

        for r in range(prima, ultima + 1):
            nome = _testo(ws.cell(r, col_nome))
            if _e_segnaposto(nome):
                continue

            sigla = _testo(ws.cell(r, col_sigla))
            if not sigla or _e_segnaposto(sigla):
                avvisi.append(
                    f'Il turno "{nome}" non ha una sigla di gruppo: non so in '
                    f'che struttura metterlo, l\'ho saltato.'
                )
                continue

            righe.append({
                'nome': nome,
                'sigla': sigla,
                'fascia': blocco['fascia'],
                'tipologia': _testo(ws.cell(r, col_nome + SCARTO_COLONNA_TIPOLOGIA)),
            })

    return righe, avvisi


def _leggi_turni(wb, ws):
    """
    I turni del foglio, ciascuno con la struttura a cui appartiene.

    Args:
        wb: cartella di lavoro.
        ws: foglio Tabelle.

    Returns:
        tuple: (lista di turni, lista di avvisi).
    """
    righe, avvisi = _righe_turni(wb, ws)
    strutture = _strutture_dei_posti([(r['sigla'], r['nome']) for r in righe])

    turni = []
    for r in righe:
        posto = _chiave_sigla(r['sigla'])
        chiave, nome_struttura = strutture[posto]
        turni.append({
            'nome': r['nome'],
            'fascia': r['fascia'],
            'posto': posto,
            'struttura': chiave,
            'struttura_nome': nome_struttura,
            'tipologia': r['tipologia'],
            'ordine': len(turni),
        })

    return turni, avvisi


def _leggi_persone(wb, ws):
    """Le persone del foglio, saltando i segnaposto."""
    colonna, prima, ultima = _intervallo(wb, NOME_PERSONE, RIPIEGO_PERSONE)

    persone = []
    for r in range(prima, ultima + 1):
        sigla = _testo(ws.cell(r, colonna))
        cognome = _testo(ws.cell(r, colonna + SCARTO_COGNOME))
        if not sigla or _e_segnaposto(sigla) or _e_segnaposto(cognome):
            continue
        if sigla.upper().startswith('VUOTO'):
            continue

        persone.append({
            'sigla': sigla.upper(),
            'cognome': cognome,
            'nome': _testo(ws.cell(r, colonna + SCARTO_NOME)),
        })

    return persone


def _sigle_in(wb, ws, nome, ripiego):
    """Le sigle non vuote di un intervallo, nell'ordine del foglio."""
    colonna, prima, ultima = _intervallo(wb, nome, ripiego)

    sigle = []
    for r in range(prima, ultima + 1):
        sigla = _testo(ws.cell(r, colonna))
        if sigla and not _e_segnaposto(sigla) and sigla not in sigle:
            sigle.append(sigla)

    return sigle


def _leggi_tipi_richiesta(wb, ws):
    """
    Le sigle con cui il foglio esprime desiderata e assenze.

    Il foglio tiene due elenchi: tutto cio' che si puo' chiedere, e quali di
    quelle cose sono un'assenza. Il secondo e' un sottoinsieme del primo, non
    una lista a parte.
    """
    assenze = {s.upper() for s in _sigle_in(wb, ws, NOME_ASSENZE, RIPIEGO_ASSENZE)}

    return [
        {'sigla': sigla,
         'tipo': TIPO_ASSENZA if sigla.upper() in assenze else TIPO_LAVORATIVO}
        for sigla in _sigle_in(wb, ws, NOME_RICHIESTE, RIPIEGO_RICHIESTE)
    ]


def rinomina_strutture(letto, rinomina):
    """
    Applica le correzioni dell'utente alle strutture dedotte.

    La sede ricavata dal nome e' una deduzione, e chi importa deve poterla
    correggere: rinominare due strutture allo stesso modo **le fonde**, ed e'
    il modo per dire «queste due sono lo stesso posto».

    Args:
        letto (dict): esito di leggi_struttura(), modificato sul posto.
        rinomina (dict): chiave della struttura → nome scelto. Le chiavi
                         assenti restano come sono.

    Returns:
        dict: lo stesso `letto`, con strutture e turni riallineati.
    """
    if not rinomina:
        return letto

    def scelto(chiave, corrente):
        nuovo_nome = (rinomina.get(chiave) or '').strip()
        return nuovo_nome or corrente

    for turno in letto['turni']:
        nome = scelto(turno['struttura'], turno['struttura_nome'])
        turno['struttura'] = nome
        turno['struttura_nome'] = nome

    strutture = []
    viste = set()
    for turno in letto['turni']:
        if turno['struttura'] not in viste:
            viste.add(turno['struttura'])
            strutture.append({
                'chiave': turno['struttura'], 'nome': turno['struttura_nome'],
            })
    letto['strutture'] = strutture

    return letto


def leggi_struttura(sorgente):
    """
    Legge da un modello Excel tutto cio' che serve a costruire la struttura.

    Args:
        sorgente: percorso, oggetto file o BytesIO del foglio.

    Returns:
        dict: strutture, turni, tipologie, persone, tipi_richiesta, avvisi.

    Raises:
        ValueError: il foglio non ha la tabella da cui si legge la struttura.
    """
    wb = openpyxl.load_workbook(sorgente, data_only=True)

    if FOGLIO_TABELLE not in wb.sheetnames:
        raise ValueError(
            f'Nel file manca il foglio "{FOGLIO_TABELLE}", che e\' quello da '
            f'cui si legge la struttura.'
        )

    ws = wb[FOGLIO_TABELLE]
    turni, avvisi = _leggi_turni(wb, ws)

    if not turni:
        raise ValueError('Nel foglio non ho trovato nessun turno.')

    # Le strutture, nell'ordine in cui i turni le incontrano.
    strutture = []
    viste = set()
    for t in turni:
        if t['struttura'] not in viste:
            viste.add(t['struttura'])
            strutture.append({'chiave': t['struttura'], 'nome': t['struttura_nome']})

    tipologie = []
    for t in turni:
        if t['tipologia'] and not _e_segnaposto(t['tipologia']) \
                and t['tipologia'] not in tipologie:
            tipologie.append(t['tipologia'])

    return {
        'strutture': strutture,
        'turni': turni,
        'tipologie': tipologie,
        'persone': _leggi_persone(wb, ws),
        'tipi_richiesta': _leggi_tipi_richiesta(wb, ws),
        'avvisi': avvisi,
    }
