"""
app/services/modello_desiderata.py — i desiderata letti da un foglio Excel.

Nuova feature. I lavoratori compilano le proprie richieste in un foglio di
calcolo — una riga per persona, una colonna per giorno — e chi pianifica se
lo ritrova pieno. Ridigitarlo nella griglia e' un lavoro lungo su dati che
esistono gia'.

Come e' fatto il foglio:

    riga 8      le date del mese, una per colonna; piu' in la', due celle
                intestate MESE e ANNO con il valore nella riga sotto
    riga 9      i nomi dei giorni della settimana
    righe 10+   una per lavoratore: la sigla in colonna A, e in ogni colonna
                la sigla della richiesta per quel giorno

Il mese non si indovina dal nome del foglio: si legge dalle celle che lo
dichiarano, e in mancanza di quelle dalle date della riga 8. Le colonne dei
giorni si ricavano dalle date, non contate dalla prima: un foglio che comincia
altrove viene letto lo stesso.

Questo modulo **non tocca il database**: legge e descrive. Confrontare con le
persone del programma e decidere cosa scrivere e' compito di chi lo chiama.
"""

import datetime

import openpyxl


# Il foglio delle tabelle di servizio: e' l'unico che non contiene un mese.
FOGLIO_TABELLE = 'TABELLE'

# Dove stanno le date e da dove cominciano i lavoratori.
RIGA_DATE = 8
PRIMA_RIGA_LAVORATORI = 10

# Quante righe scorrere in cerca di lavoratori prima di arrendersi: i fogli
# veri ne hanno una quarantina, ma sotto c'e' un piede di pagina da non
# scambiare per una persona.
RIGHE_DA_SCORRERE = 80

# Le intestazioni che dichiarano il mese e l'anno, con il valore nella riga
# sotto. Si cercano nella riga delle date, dove il modello le tiene.
INTESTAZIONE_MESE = 'MESE'
INTESTAZIONE_ANNO = 'ANNO'

# I segnaposto con cui il foglio riempie le posizioni non coperte.
SEGNAPOSTO = ('vuoto', 'chiusa', 'chiuso', '-', '')

# Prefisso dei segnaposto numerati: VUOTO_C, VUOTO_D, ...
PREFISSO_SEGNAPOSTO = 'VUOTO'


def _testo(cella):
    """Il contenuto di una cella come testo pulito, o stringa vuota."""
    if cella is None or cella.value is None:
        return ''

    return str(cella.value).strip()


def _e_segnaposto(testo):
    """La riga e' una di quelle che il foglio tiene libere."""
    pulito = testo.strip()

    return (pulito.lower() in SEGNAPOSTO
            or pulito.upper().startswith(PREFISSO_SEGNAPOSTO))


def _foglio_del_mese(wb):
    """
    Il foglio con la griglia dei desiderata.

    Args:
        wb: cartella di lavoro.

    Returns:
        worksheet: il primo foglio che non sia quello delle tabelle.

    Raises:
        ValueError: la cartella ha solo tabelle di servizio.
    """
    for nome in wb.sheetnames:
        if nome.strip().upper() != FOGLIO_TABELLE:
            return wb[nome]

    raise ValueError('Nel file non c\'e\' nessun foglio con i desiderata.')


def _colonne_dei_giorni(ws):
    """
    Quale colonna corrisponde a quale giorno del mese.

    Si ricava dalle date scritte nella riga delle intestazioni: contarle dalla
    prima colonna darebbe per scontato che il foglio cominci sempre allo
    stesso punto.

    Args:
        ws: foglio con la griglia.

    Returns:
        dict: indice di colonna → giorno del mese.
    """
    giorni = {}
    for c in range(1, ws.max_column + 1):
        valore = ws.cell(RIGA_DATE, c).value
        if isinstance(valore, datetime.datetime):
            giorni[c] = valore.day
        elif isinstance(valore, datetime.date):
            giorni[c] = valore.day

    return giorni


def _mese_e_anno(ws, colonne_giorni):
    """
    Di che mese parla il foglio.

    Lo dicono due celle intestate MESE e ANNO; se mancano, lo dicono le date
    stesse. Fidarsi del nome del foglio sarebbe peggio: si rinomina.

    Args:
        ws: foglio con la griglia.
        colonne_giorni (dict): colonne che portano una data.

    Returns:
        tuple: (mese, anno).

    Raises:
        ValueError: il foglio non dice di che mese parla.
    """
    dichiarati = {}
    for c in range(1, ws.max_column + 1):
        etichetta = _testo(ws.cell(RIGA_DATE, c)).upper()
        if etichetta in (INTESTAZIONE_MESE, INTESTAZIONE_ANNO):
            valore = ws.cell(RIGA_DATE + 1, c).value
            try:
                dichiarati[etichetta] = int(valore)
            except (TypeError, ValueError):
                pass

    if INTESTAZIONE_MESE in dichiarati and INTESTAZIONE_ANNO in dichiarati:
        return dichiarati[INTESTAZIONE_MESE], dichiarati[INTESTAZIONE_ANNO]

    for c in colonne_giorni:
        data = ws.cell(RIGA_DATE, c).value
        if isinstance(data, (datetime.date, datetime.datetime)):
            return data.month, data.year

    raise ValueError(
        'Il foglio non dice di che mese e anno parla: servono le celle MESE e '
        'ANNO, oppure le date nella riga delle intestazioni.'
    )


def _righe_dei_lavoratori(ws):
    """
    Le righe con una persona, dalla prima fino al primo buco.

    Il buco e' il confine: sotto ci sono piedi di pagina e tabelle di
    servizio, che non sono lavoratori.

    Args:
        ws: foglio con la griglia.

    Returns:
        list: [(riga, sigla)] delle sole persone vere.
    """
    righe = []
    for r in range(PRIMA_RIGA_LAVORATORI, PRIMA_RIGA_LAVORATORI + RIGHE_DA_SCORRERE):
        sigla = _testo(ws.cell(r, 1))
        if not sigla:
            break
        if _e_segnaposto(sigla):
            continue
        righe.append((r, sigla.upper()))

    return righe


def leggi_desiderata(sorgente):
    """
    Legge da un foglio Excel le richieste di un mese.

    Args:
        sorgente: percorso, oggetto file o BytesIO del foglio.

    Returns:
        dict:
            - mese, anno (int)
            - foglio (str): nome del foglio letto
            - richieste (list): [{sigla, giorno, codice}]
            - persone (list): sigle trovate, nell'ordine del foglio
            - codici (list): sigle di richiesta usate, senza ripetizioni
            - avvisi (list[str])

    Raises:
        ValueError: il foglio non e' leggibile come modello desiderata.
    """
    wb = openpyxl.load_workbook(sorgente, data_only=True)
    ws = _foglio_del_mese(wb)

    colonne_giorni = _colonne_dei_giorni(ws)
    if not colonne_giorni:
        raise ValueError(
            f'Nel foglio "{ws.title}" la riga {RIGA_DATE} non contiene date: '
            f'non so quale colonna sia quale giorno.'
        )

    mese, anno = _mese_e_anno(ws, colonne_giorni)
    lavoratori = _righe_dei_lavoratori(ws)
    if not lavoratori:
        raise ValueError('Nel foglio non ho trovato nessun lavoratore.')

    richieste = []
    codici = []
    for riga, sigla in lavoratori:
        for colonna, giorno in colonne_giorni.items():
            codice = _testo(ws.cell(riga, colonna))
            if not codice:
                continue
            richieste.append({'sigla': sigla, 'giorno': giorno, 'codice': codice})
            if codice not in codici:
                codici.append(codice)

    avvisi = []
    fuori_mese = [r for r in richieste if not 1 <= r['giorno'] <= 31]
    if fuori_mese:
        avvisi.append(f'{len(fuori_mese)} richieste cadono fuori dal mese: le salto.')
        richieste = [r for r in richieste if 1 <= r['giorno'] <= 31]

    return {
        'mese': mese,
        'anno': anno,
        'foglio': ws.title,
        'richieste': richieste,
        'persone': [sigla for _, sigla in lavoratori],
        'codici': codici,
        'avvisi': avvisi,
    }
