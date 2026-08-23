"""
app/routes/export.py — route per l'export dei dati in formato Excel.

Endpoint:
    GET /api/export/calendari/<id>/turni   → griglia turni del mese (xlsx)
    GET /api/export/calendari/<id>/ore     → riepilogo ore mensili (xlsx)
    GET /api/export/annuale/<anno>/ore     → statistiche ore annuali (xlsx)
"""

import io
import json
import calendar as cal_lib

from flask import Blueprint, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.auth import require_role
from app.db import query_one, query_all, execute_write
from app.services.ore import calcola_ore_mensili, calcola_statistiche_annuali
from app.services.validatori import valida_assegnazione
from app.services.config_snapshot import carica_appearance_snapshot, APPEARANCE_DEFAULT

bp = Blueprint('export', __name__, url_prefix='/api/export')

# ---------------------------------------------------------------------------
# Stili comuni
# ---------------------------------------------------------------------------

_THIN = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_HEADER   = PatternFill('solid', fgColor='1F4E79')   # blu scuro
_FILL_GIORNO   = PatternFill('solid', fgColor='D6E4F0')   # azzurro chiaro
_FILL_FESTIVO  = PatternFill('solid', fgColor='FCE4D6')   # arancio chiaro
_FILL_SUPER    = PatternFill('solid', fgColor='F4CCCC')   # rosso chiaro
_FILL_MATCH    = PatternFill('solid', fgColor='D9EAD3')   # verde chiaro
_FILL_MISMATCH = PatternFill('solid', fgColor='FFF2CC')   # giallo chiaro
_FILL_FORCED   = PatternFill('solid', fgColor='FCE5CD')   # arancio tenue
_FILL_EMPTY    = PatternFill('solid', fgColor='F3F3F3')   # grigio chiaro

_FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
_FONT_BOLD   = Font(bold=True)
_ALIGN_C     = Alignment(horizontal='center', vertical='center', wrap_text=True)
_ALIGN_L     = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _css_to_hex(css_color, default='FFFFFF'):
    """Converte un colore CSS (#RRGGBB o RRGGBB) al formato hex openpyxl (6 char)."""
    raw = str(css_color or default).strip().lstrip('#')
    return raw if len(raw) == 6 else default


def _fill_from_css(css_color, default='FFFFFF'):
    """Crea PatternFill da un colore CSS."""
    return PatternFill('solid', fgColor=_css_to_hex(css_color, default))


def _cell(ws, row, col, value='', fill=None, font=None, align=None, border=True):
    """Scrive una cella con stili opzionali."""
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    if font:
        c.font = font
    if align:
        c.alignment = align
    if border:
        c.border = _BORDER
    return c


def _response_xlsx(wb, filename):
    """Serializza il workbook e lo restituisce come risposta HTTP."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# =============================================================================
# EXPORT GRIGLIA TURNI MENSILE
# =============================================================================

@bp.route('/calendari/<int:cal_id>/turni', methods=['GET'])
@require_role('admin', 'manager')
def export_turni(cal_id):
    """
    Genera un file Excel con la griglia delle assegnazioni turni del mese.

    Struttura:
        - Una riga per ogni turno (raggruppati per sovragruppo/gruppo).
        - Una colonna per ogni giorno del mese.
        - Celle colorate in base al codice conflitto dell'assegnazione.

    Args:
        cal_id (int): ID del calendario.

    Returns:
        200: file xlsx in allegato.
        404: { ok: false, errore: str }
    """
    cal = query_one(
        "SELECT id, mese, anno, stato, tipo, style FROM calendari WHERE id=?", (cal_id,)
    )
    if not cal:
        return jsonify({'ok': False, 'errore': 'Calendario non trovato.'}), 404
    if cal['stato'] == 'APERTO' and cal.get('tipo', 'programmato') != 'effettivo':
        return jsonify({'ok': False, 'errore': 'Export disponibile solo per calendari chiusi.'}), 400

    # Carica appearance dal snapshot (con fallback ai default)
    app = carica_appearance_snapshot(cal_id)
    fill_festivo   = _fill_from_css(app.get('festivi_bg',       APPEARANCE_DEFAULT['festivi_bg']))
    fill_super     = _fill_from_css(app.get('superfestivi_bg',  APPEARANCE_DEFAULT['superfestivi_bg']))
    fill_prima_riga = _fill_from_css(app.get('prima_riga_bg',   APPEARANCE_DEFAULT['prima_riga_bg']))

    mese = cal['mese']
    anno = cal['anno']
    num_giorni = cal_lib.monthrange(anno, mese)[1]

    giorni_cal = {
        g['giorno']: g for g in query_all(
            "SELECT giorno, is_lavorativo, tipo "
            "FROM giorni_calendario WHERE calendario_id=? ORDER BY giorno",
            (cal_id,)
        )
    }

    # Ricalcola conflitti per tutte le assegnazioni assegnate
    # (garantisce dati aggiornati anche per calendari pre-snapshot)
    ass_rows = query_all(
        "SELECT id, turno_id, giorno, user_id "
        "FROM assegnazioni_turni WHERE calendario_id=? AND user_id IS NOT NULL",
        (cal_id,)
    )
    for a in ass_rows:
        try:
            res = valida_assegnazione(cal_id, a['turno_id'], a['user_id'], a['giorno'])
            execute_write(
                "UPDATE assegnazioni_turni SET conflitti=? WHERE id=?",
                (json.dumps(res['conflitti']), a['id'])
            )
        except Exception:
            pass

    # Leggi assegnazioni con conflitti aggiornati
    assegnazioni = {
        (a['turno_id'], a['giorno']): a
        for a in query_all(
            "SELECT turno_id, giorno, user_id, conflitti "
            "FROM assegnazioni_turni WHERE calendario_id=?",
            (cal_id,)
        )
    }

    # Mappa sigla utente
    utenti = {u['id']: u['sigla'] for u in query_all(
        "SELECT id, sigla FROM users", ()
    )}

    # Struttura gerarchica turni dallo snapshot calendario_turni (include style)
    cal_style = json.loads(cal.get('style', '{}') or '{}')
    ct_rows = query_all(
        "SELECT id, nome, sigla, sg_sigla, sg_nome, sg_ordine, sg_style, "
        "gruppo_sigla, gruppo_nome, gruppo_ordine, style, ordine "
        "FROM calendario_turni WHERE calendario_id=? "
        "ORDER BY sg_ordine, gruppo_ordine, ordine",
        (cal_id,)
    )
    sovragruppi = []
    sg_map = {}
    for row in ct_rows:
        sg_key = row['sg_sigla']
        if sg_key not in sg_map:
            sg_style = json.loads(row.get('sg_style', '{}') or '{}')
            sg_map[sg_key] = {'sigla': row['sg_sigla'], 'nome': row['sg_nome'],
                              'style': sg_style, '_gruppi': {}}
            sovragruppi.append(sg_map[sg_key])
        g_key = (row['gruppo_sigla'], row['gruppo_ordine'])
        g_map = sg_map[sg_key]['_gruppi']
        if g_key not in g_map:
            g_style = json.loads(row.get('style', '{}') or '{}')
            g_map[g_key] = {
                'sigla': row['gruppo_sigla'], 'nome': row['gruppo_nome'],
                'style': {**cal_style, **g_style}, 'turni': []
            }
        g_map[g_key]['turni'].append({
            'id': row['id'],
            'sigla': row['sigla'],
            'nome': row.get('nome') or row['sigla'],
        })
    for sg in sovragruppi:
        sg['gruppi'] = list(sg.pop('_gruppi').values())

    # ---------------------------------------------------------------------------
    # Costruzione foglio
    # ---------------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = f'{mese:02d}-{anno}'

    NOME_MESI = [
        '', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
        'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre',
    ]
    NOME_GIORNI = ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom']

    # Colori header riga 1 (personalizzabili via query param)
    header_bg_raw = request.args.get('header_bg', '1F4E79').lstrip('#')
    header_fg_raw = request.args.get('header_fg', 'FFFFFF').lstrip('#')
    fill_header_custom = PatternFill('solid', fgColor=header_bg_raw)
    font_header_custom = Font(bold=True, color=header_fg_raw, size=11)

    # Riga 1: titolo
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=1 + num_giorni)
    _cell(ws, 1, 1,
          f'Turni {NOME_MESI[mese]} {anno}  —  {cal["stato"]}',
          fill=fill_header_custom, font=font_header_custom, align=_ALIGN_C)
    ws.row_dimensions[1].height = 22

    # Larghezza prima colonna dal style del calendario
    col_turno_width = max(15, cal_style.get('--colTurnoWidth', 155) / 7)

    # Riga 2: intestazioni giorni (numero + giorno settimana)
    _cell(ws, 2, 1, 'Turno', fill=fill_prima_riga, font=_FONT_BOLD, align=_ALIGN_C)
    ws.column_dimensions['A'].width = col_turno_width

    import datetime
    for g in range(1, num_giorni + 1):
        col = g + 1
        data = datetime.date(anno, mese, g)
        nome_g = NOME_GIORNI[data.weekday()]
        gg = giorni_cal.get(g, {})
        tipo = gg.get('tipo', 'normale')
        intestazione_fill = (
            fill_super      if tipo == 'superfestivo' else
            fill_festivo    if tipo == 'festivo'      else
            fill_prima_riga
        )
        _cell(ws, 2, col, f'{g}\n{nome_g}',
              fill=intestazione_fill, font=_FONT_BOLD, align=_ALIGN_C)
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.row_dimensions[2].height = 28

    # Righe dati: una per turno
    riga = 3
    def _fill_da_conflitti(conflitti_json):
        """
        Restituisce PatternFill e Font dal conflitto a peso più alto.
        Formato conflitti: [{id, nome, categoria, stile (JSON), peso_numerico, ...}]
        """
        try:
            items = json.loads(conflitti_json) if isinstance(conflitti_json, str) else (conflitti_json or [])
        except (json.JSONDecodeError, TypeError):
            items = []
        if not items:
            return None, None
        # Seleziona il conflitto con peso maggiore
        best = max(items, key=lambda c: c.get('peso_numerico', 0) if isinstance(c, dict) else 0)
        if not isinstance(best, dict):
            return None, None
        # Estrai stile dal campo 'stile' (JSON string o dict)
        try:
            stile = json.loads(best['stile']) if isinstance(best.get('stile'), str) else (best.get('stile') or {})
        except (json.JSONDecodeError, TypeError):
            stile = {}
        bg = stile.get('backgroundColor', '#ffffff').lstrip('#')
        fg = stile.get('color', '#000000').lstrip('#')
        bold = stile.get('fontWeight', '') == 'bold'
        italic = stile.get('fontStyle', '') == 'italic'
        fill = PatternFill('solid', fgColor=bg)
        font = Font(color=fg, bold=bold, italic=italic)
        return fill, font

    max_cell_len = 3  # per auto-width colonne giorni

    for sg in sovragruppi:
        # Riga separatore sovragruppo con formattazione personalizzata
        sg_st = sg.get('style', {})
        sg_bg = sg_st.get('backgroundColor', '#d6d8db').lstrip('#')
        sg_fg = sg_st.get('color', '#1a1a1a').lstrip('#')
        sg_font_name = sg_st.get('fontFamily', '')
        sg_bold = sg_st.get('fontWeight', 'bold') == 'bold'
        sg_repeat = sg_st.get('--repeatName', False)

        sg_fill = PatternFill('solid', fgColor=sg_bg)
        sg_font = Font(bold=sg_bold, color=sg_fg, size=10,
                       name=sg_font_name or None)

        if sg_repeat:
            # Col 1 (label): solo fill
            _cell(ws, riga, 1, '', fill=sg_fill)
            # 3 blocchi di celle unite simmetrici sulle colonne giorni (2..N+1)
            g1 = num_giorni // 3
            g2 = num_giorni // 3
            g3 = num_giorni - g1 - g2
            blocchi = [
                (2, 2 + g1 - 1),
                (2 + g1, 2 + g1 + g2 - 1),
                (2 + g1 + g2, 1 + num_giorni),
            ]
            nome_sg = sg.get('nome', sg['sigla'])
            for c_start, c_end in blocchi:
                if c_start <= c_end:
                    ws.merge_cells(
                        start_row=riga, start_column=c_start,
                        end_row=riga, end_column=c_end
                    )
                    _cell(ws, riga, c_start, nome_sg,
                          fill=sg_fill, font=sg_font, align=_ALIGN_C)
        else:
            ws.merge_cells(start_row=riga, start_column=1,
                           end_row=riga, end_column=1 + num_giorni)
            _cell(ws, riga, 1, sg.get('nome', sg['sigla']),
                  fill=sg_fill, font=sg_font, align=_ALIGN_L)
        ws.row_dimensions[riga].height = 18
        riga += 1

        for g in sg['gruppi']:
            # Riga separatore gruppo con formattazione personalizzata
            g_style = g.get('style', {})
            bg = g_style.get('backgroundColor', '#e9ecef').lstrip('#')
            fg = g_style.get('color', '#6c757d').lstrip('#')
            font_name = g_style.get('fontFamily', '')
            font_bold = g_style.get('fontWeight', '') == 'bold'
            font_italic = g_style.get('fontStyle', '') == 'italic'
            font_size = None
            fs_raw = g_style.get('fontSize', '')
            if fs_raw and fs_raw.endswith('rem'):
                try:
                    font_size = max(7, int(float(fs_raw.replace('rem', '')) * 11))
                except ValueError:
                    pass
            repeat_name = g_style.get('--repeatName', False)

            sep_fill = PatternFill('solid', fgColor=bg)
            sep_font = Font(
                bold=font_bold, italic=font_italic, color=fg,
                size=font_size or 9,
                name=font_name or None,
            )

            if repeat_name:
                # Col 1 (label): solo fill
                _cell(ws, riga, 1, '', fill=sep_fill)
                # 3 blocchi di celle unite simmetrici sulle colonne giorni (2..N+1)
                g1 = num_giorni // 3
                g2 = num_giorni // 3
                blocchi_g = [
                    (2, 2 + g1 - 1),
                    (2 + g1, 2 + g1 + g2 - 1),
                    (2 + g1 + g2, 1 + num_giorni),
                ]
                nome_g = g.get('nome', g['sigla'])
                for c_start, c_end in blocchi_g:
                    if c_start <= c_end:
                        ws.merge_cells(
                            start_row=riga, start_column=c_start,
                            end_row=riga, end_column=c_end
                        )
                        _cell(ws, riga, c_start, nome_g,
                              fill=sep_fill, font=sep_font, align=_ALIGN_C)
            else:
                ws.merge_cells(start_row=riga, start_column=1,
                               end_row=riga, end_column=1 + num_giorni)
                _cell(ws, riga, 1, g.get('nome', g['sigla']),
                      fill=sep_fill, font=sep_font, align=_ALIGN_L)
            ws.row_dimensions[riga].height = 16
            riga += 1

            for t in g['turni']:
                label = t['nome']
                _cell(ws, riga, 1, label, font=_FONT_BOLD, align=_ALIGN_L)

                for giorno in range(1, num_giorni + 1):
                    col = giorno + 1
                    ass = assegnazioni.get((t['id'], giorno))
                    if ass:
                        testo = utenti.get(ass['user_id'], '—') if ass['user_id'] else '—'
                        fill, font_conf = _fill_da_conflitti(ass.get('conflitti', '[]'))
                        if testo and len(testo) > max_cell_len:
                            max_cell_len = len(testo)
                    else:
                        testo = ''
                        fill = None
                        font_conf = None
                    _cell(ws, riga, col, testo, fill=fill,
                          font=font_conf, align=_ALIGN_C)

                ws.row_dimensions[riga].height = 16
                riga += 1

    # Auto-width colonne giorni in base al contenuto più lungo
    auto_w = max(6, max_cell_len * 1.3 + 1)
    for g in range(1, num_giorni + 1):
        ws.column_dimensions[get_column_letter(g + 1)].width = auto_w

    # Blocca intestazioni
    ws.freeze_panes = 'B3'

    return _response_xlsx(wb, f'turni_{mese:02d}_{anno}.xlsx')


# =============================================================================
# EXPORT ORE MENSILI
# =============================================================================

@bp.route('/calendari/<int:cal_id>/ore', methods=['GET'])
@require_role('admin', 'manager')
def export_ore_mensili(cal_id):
    """
    Genera un file Excel con il riepilogo ore mensile per tutti i lavoratori.

    Args:
        cal_id (int): ID del calendario.

    Returns:
        200: file xlsx in allegato.
        404: { ok: false, errore: str }
    """
    cal = query_one("SELECT id, mese, anno, stato, tipo FROM calendari WHERE id=?", (cal_id,))
    if not cal:
        return jsonify({'ok': False, 'errore': 'Calendario non trovato.'}), 404
    if cal['stato'] == 'APERTO' and cal.get('tipo', 'programmato') != 'effettivo':
        return jsonify({'ok': False, 'errore': 'Export disponibile solo per calendari chiusi.'}), 400

    ore = calcola_ore_mensili(cal_id)
    mese = cal['mese']
    anno = cal['anno']

    wb = Workbook()
    ws = wb.active
    ws.title = f'Ore {mese:02d}-{anno}'

    intestazioni = [
        'Sigla', 'Ore Lavorate', 'Ore Giustificate', 'Ore Totali',
        'Gg Lavorati', 'Gg Giustificati', 'Gg Festivi Lavorati', 'Gg Superfestivi',
    ]
    chiavi = [
        'sigla', 'ore_lavorate', 'ore_giustificate', 'ore_totali',
        'giorni_lavorati', 'giorni_giustificati',
        'giorni_festivi_lavorati', 'giorni_superfestivi_lavorati',
    ]

    # Titolo
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(intestazioni))
    _cell(ws, 1, 1, f'Ore mensili — {mese:02d}/{anno}',
          fill=_FILL_HEADER, font=_FONT_HEADER, align=_ALIGN_C)
    ws.row_dimensions[1].height = 22

    # Intestazioni colonne
    for col, nome in enumerate(intestazioni, start=1):
        _cell(ws, 2, col, nome,
              fill=_FILL_GIORNO, font=_FONT_BOLD, align=_ALIGN_C)
    ws.row_dimensions[2].height = 20

    # Dati
    for riga, lav in enumerate(ore, start=3):
        for col, chiave in enumerate(chiavi, start=1):
            _cell(ws, riga, col, lav.get(chiave, ''), align=_ALIGN_C)

    # Larghezze colonne
    larghezze = [12, 14, 18, 12, 12, 16, 20, 16]
    for col, w in enumerate(larghezze, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return _response_xlsx(wb, f'ore_{mese:02d}_{anno}.xlsx')


# =============================================================================
# EXPORT STATISTICHE ANNUALI
# =============================================================================

@bp.route('/annuale/<int:anno>/ore', methods=['GET'])
@require_role('admin', 'manager')
def export_ore_annuali(anno):
    """
    Genera un file Excel con le statistiche ore annuali aggregate per lavoratore.

    Include un foglio riepilogativo e un foglio di dettaglio mensile per ogni
    lavoratore.

    Args:
        anno (int): anno di riferimento.

    Returns:
        200: file xlsx in allegato.
    """
    escludi_param = request.args.get('escludi', '')
    escludi_ids = []
    if escludi_param:
        escludi_ids = [int(x) for x in escludi_param.split(',') if x.strip().isdigit()]

    statistiche = calcola_statistiche_annuali(anno, escludi_ids=escludi_ids)

    wb = Workbook()

    # -----------------------------------------------------------------------
    # Foglio 1: riepilogo annuale
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = f'Riepilogo {anno}'

    intestazioni = [
        'Sigla', 'Ore Totali Anno',
        'Gg Festivi Lavorati', 'Gg Superfestivi Lavorati',
    ]
    chiavi = [
        'sigla', 'ore_totali_anno',
        'giorni_festivi_anno', 'giorni_superfestivi_anno',
    ]

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(intestazioni))
    _cell(ws, 1, 1, f'Statistiche annuali {anno}',
          fill=_FILL_HEADER, font=_FONT_HEADER, align=_ALIGN_C)
    ws.row_dimensions[1].height = 22

    for col, nome in enumerate(intestazioni, start=1):
        _cell(ws, 2, col, nome,
              fill=_FILL_GIORNO, font=_FONT_BOLD, align=_ALIGN_C)
    ws.row_dimensions[2].height = 20

    for riga, lav in enumerate(statistiche, start=3):
        for col, chiave in enumerate(chiavi, start=1):
            _cell(ws, riga, col, lav.get(chiave, ''), align=_ALIGN_C)

    for col, w in enumerate([12, 18, 22, 22], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # -----------------------------------------------------------------------
    # Foglio 2: dettaglio mensile per lavoratore
    # -----------------------------------------------------------------------
    NOME_MESI = [
        '', 'Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu',
        'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic',
    ]

    ws2 = wb.create_sheet(title=f'Dettaglio {anno}')

    # Raccoglie tutti i mesi presenti
    mesi_presenti = sorted({
        m['mese']
        for lav in statistiche
        for m in lav.get('mesi_dettaglio', [])
    })

    header_row = ['Sigla'] + [NOME_MESI[m] for m in mesi_presenti] + ['Totale']
    for col, nome in enumerate(header_row, start=1):
        _cell(ws2, 1, col, nome,
              fill=_FILL_HEADER, font=_FONT_HEADER, align=_ALIGN_C)
    ws2.row_dimensions[1].height = 20

    for riga, lav in enumerate(statistiche, start=2):
        mese_map = {m['mese']: m['ore_totali'] for m in lav.get('mesi_dettaglio', [])}
        _cell(ws2, riga, 1, lav['sigla'], font=_FONT_BOLD, align=_ALIGN_C)
        for col, mese in enumerate(mesi_presenti, start=2):
            _cell(ws2, riga, col, mese_map.get(mese, 0), align=_ALIGN_C)
        _cell(ws2, riga, len(mesi_presenti) + 2,
              lav['ore_totali_anno'], font=_FONT_BOLD, align=_ALIGN_C)

    for col in range(1, len(header_row) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 9
    ws2.column_dimensions['A'].width = 12

    return _response_xlsx(wb, f'ore_annuali_{anno}.xlsx')
